"""
Exports Excel et PDF pour les bons de commande (achat).
"""
from io import BytesIO


def branche_siege_pour(entreprise):
    from entreprise.models import Branche

    return (
        Branche.objects.filter(entreprise=entreprise, est_siege_social=True)
        .order_by('pk')
        .first()
    )


def _esc_pdf(txt):
    if txt is None:
        return ''
    s = str(txt)
    return s.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')


def build_excel_bc(commande):
    """Retourne le contenu binaire d'un classeur .xlsx (récap + lignes)."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = 'Récap'

    ent = commande.entreprise
    four = commande.fournisseur
    bs = branche_siege_pour(ent)

    titre_font = Font(bold=True, size=14)
    ws['A1'] = f'Bon de commande {commande.numero_commande}'
    ws['A1'].font = titre_font
    ws.merge_cells('A1:B1')

    meta = [
        ('Entreprise', ent.nom),
        ('Adresse siège social', ent.adresse_siege or ''),
        ('RCCM', ent.rccm or ''),
        ('ID National', ent.idnat or ''),
        ('N° impôt', ent.numero_impot or ''),
        ('Téléphone', ent.telephone or ''),
        ('Email', ent.email or ''),
        ('Branche siège social', bs.nom if bs else ''),
        ('Ville (siège)', bs.ville if bs else ''),
        ('Code branche siège', bs.code_branche if bs else ''),
        ('', ''),
        ('Fournisseur', four.nom_societe),
        ('Code fournisseur', four.code_fournisseur),
        ('Contact', four.contact_nom or ''),
        ('Tél. fournisseur', four.telephone or ''),
        ('Email fournisseur', four.email or ''),
        ('Adresse fournisseur', four.adresse or ''),
        ('Ville fournisseur', four.ville or ''),
        ('ID Fiscal / RCCM (tiers)', four.rccm_id or ''),
        ('', ''),
        ('N° commande', commande.numero_commande),
        ('Date commande', commande.date_commande.strftime('%d/%m/%Y %H:%M')),
        ('Statut', commande.get_statut_display()),
        ('Livraison prévue', commande.date_livraison_prevue.strftime('%d/%m/%Y') if commande.date_livraison_prevue else ''),
        ('Dépôt destination', commande.depot_destination.nom if commande.depot_destination_id else ''),
        ('Point de vente destination', commande.pointdevente_destination.nom if commande.pointdevente_destination_id else ''),
        ('Devise', commande.devise.code if commande.devise_id else ''),
        ('Créé par', commande.cree_par.get_username() if commande.cree_par_id else ''),
        ('Notes commande', commande.notes or ''),
        ('', ''),
        ('Total HT', float(commande.total_ht)),
        ('Total TVA', float(commande.total_tva)),
        ('Total TTC', float(commande.total_ttc)),
    ]

    row = 3
    label_fill = PatternFill('solid', fgColor='E8EEF7')
    for label, val in meta:
        ws.cell(row=row, column=1, value=label)
        ws.cell(row=row, column=2, value=val)
        if label:
            ws.cell(row=row, column=1).font = Font(bold=True)
            ws.cell(row=row, column=1).fill = label_fill
        ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical='top')
        row += 1

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 55

    ws2 = wb.create_sheet('Lignes')
    headers = [
        'sku',
        'nom_produit',
        'code_barre',
        'quantite_commandee',
        'unite',
        'prix_unitaire_ht',
        'taux_tva_pct',
        'montant_ht_ligne',
        'quantite_recue',
        'lot_batch',
        'dateproduction',
        'dateexpiration',
        'location_code',
    ]
    h_font = Font(bold=True, color='FFFFFF')
    h_fill = PatternFill('solid', fgColor='1A56DB')
    for col, h in enumerate(headers, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = h_font
        c.fill = h_fill

    for ligne in commande.lignes.all():
        p = ligne.produit
        ws2.append(
            [
                (p.sku or '') if p.sku else '',
                p.nom,
                p.code_barre or '',
                float(ligne.quantite_commandee),
                ligne.unite or '',
                float(ligne.prix_unitaire_ht),
                float(p.tva_taux),
                float(ligne.sous_total_ht),
                float(ligne.quantite_recue or 0),
                ligne.lot_batch or '',
                ligne.dateproduction.isoformat() if ligne.dateproduction else '',
                ligne.dateexpiration.isoformat() if ligne.dateexpiration else '',
                ligne.location.code if ligne.location_id else '',
            ]
        )

    for col in range(1, len(headers) + 1):
        ws2.column_dimensions[ws2.cell(row=1, column=col).column_letter].width = 16

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _statut_couleur(statut):
    return {
        'BROUILLON': '#64748B',
        'ENVOYE': '#1A56DB',
        'RECU_PARTIEL': '#D97706',
        'RECU_TOTAL': '#059669',
        'ANNULE': '#DC2626',
    }.get(statut, '#64748B')


def _symbole_devise(commande):
    if commande.devise_id and getattr(commande.devise, 'symbole', None):
        return commande.devise.symbole
    return '$'


def _qr_flowable(commande):
    """QR code pointant vers la fiche BC si SITE_PUBLIC_URL est défini, sinon référence interne."""
    from io import BytesIO as _BIO

    try:
        from django.conf import settings
        from django.urls import reverse

        base = getattr(settings, 'SITE_PUBLIC_URL', '').strip()
        if base:
            txt = base.rstrip('/') + reverse('achat:detail-commande', args=[commande.pk])
        else:
            txt = f'BC:{commande.numero_commande}|PK:{commande.pk}|ENT:{commande.entreprise_id}'
    except Exception:
        txt = f'BC:{commande.numero_commande}|PK:{commande.pk}'

    import qrcode
    from reportlab.lib.units import cm
    from reportlab.platypus import Image as RLImage

    buf = _BIO()
    qr = qrcode.QRCode(version=None, box_size=3, border=1)
    qr.add_data(txt)
    qr.make(fit=True)
    img = qr.make_image(fill_color='#111827', back_color='white')
    img.save(buf, format='PNG')
    buf.seek(0)
    return RLImage(buf, width=2.6 * cm, height=2.6 * cm)


def build_pdf_bc(commande):
    """PDF type document commercial : en-tête, émetteur / fournisseur, QR, lignes, totaux, signatures."""
    import os

    from django.conf import settings

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Image as RLImage
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    buf = BytesIO()
    margin = 1.6 * cm
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=margin,
        rightMargin=margin,
        topMargin=margin,
        bottomMargin=margin,
        title=f'BC {commande.numero_commande}',
    )

    W = A4[0] - 2 * margin
    styles = getSampleStyleSheet()

    title_doc = ParagraphStyle(
        'title_doc',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=11,
        textColor=colors.HexColor('#0F172A'),
        leading=14,
    )
    muted = ParagraphStyle(
        'muted',
        parent=styles['Normal'],
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor('#475569'),
    )
    body = ParagraphStyle(
        'body',
        parent=styles['Normal'],
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#334155'),
    )
    small_white = ParagraphStyle(
        'small_white',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8,
        textColor=colors.white,
        alignment=1,
    )

    sym = _symbole_devise(commande)
    ent = commande.entreprise
    four = commande.fournisseur
    bs = branche_siege_pour(ent)

    story = []

    # --- Logo entreprise (optionnel) ---
    logo_flow = None
    try:
        if ent.logo:
            pth = ent.logo.path
            if os.path.isfile(pth):
                logo_flow = RLImage(pth)
                logo_flow.restrictSize(2.5 * cm, 1.5 * cm)
    except Exception:
        logo_flow = None

    siege_txt = ''
    if bs:
        siege_txt = (
            f'<font size="9"><b>Branche siège social</b><br/>'
            f'{_esc_pdf(bs.nom)} — {_esc_pdf(bs.ville)} (code {_esc_pdf(bs.code_branche)})</font>'
        )

    emitter_detail_html = (
        f'<font size="11" color="#0F172A"><b>{_esc_pdf(ent.nom)}</b></font><br/><br/>'
        f'<font size="9">{_esc_pdf(ent.adresse_siege or "")}<br/><br/>'
        f'Tél. {_esc_pdf(ent.telephone)} · {_esc_pdf(ent.email)}<br/>'
        f'RCCM {_esc_pdf(ent.rccm)} · ID Nat. {_esc_pdf(ent.idnat)} · N° impôt {_esc_pdf(ent.numero_impot)}'
        f'</font>'
    )
    if siege_txt:
        emitter_detail_html += '<br/><br/>' + siege_txt

    emitter_compact_html = (
        f'<font size="13" color="#0F172A"><b>{_esc_pdf(ent.nom)}</b></font><br/>'
        f'<font size="8.5" color="#64748B">{_esc_pdf((ent.adresse_siege or "")[:280])}</font>'
    )

    left_stack = []
    if logo_flow:
        left_stack.append([logo_flow])
        left_stack.append([Spacer(1, 0.15 * cm)])
    left_stack.append([Paragraph(emitter_compact_html, body)])

    left_tbl = Table(left_stack, colWidths=[W * 0.52])
    left_tbl.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP'), ('LEFTPADDING', (0, 0), (-1, -1), 0)]))

    meta_lines = [
        [Paragraph('<font size="8" color="#64748B">N° bon de commande</font>', body),
         Paragraph(f'<b>{_esc_pdf(commande.numero_commande)}</b>', title_doc)],
        [Paragraph('<font size="8" color="#64748B">Date d\'émission</font>', body),
         Paragraph(_esc_pdf(commande.date_commande.strftime('%d/%m/%Y à %H:%M')), title_doc)],
        [Paragraph('<font size="8" color="#64748B">Livraison prévue</font>', body),
         Paragraph(
             _esc_pdf(
                 commande.date_livraison_prevue.strftime('%d/%m/%Y')
                 if commande.date_livraison_prevue
                 else '—'
             ),
             title_doc,
         )],
        [Paragraph('<font size="8" color="#64748B">Devise</font>', body),
         Paragraph(_esc_pdf(commande.devise.code if commande.devise_id else '—'), title_doc)],
    ]
    meta_tbl = Table(meta_lines, colWidths=[3.4 * cm, 4.2 * cm])
    meta_tbl.setStyle(
        TableStyle(
            [
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 0), (-1, -2), 6),
            ]
        )
    )

    st_bg = colors.HexColor(_statut_couleur(commande.statut))
    statut_cell = Table(
        [[Paragraph(_esc_pdf(commande.get_statut_display()), small_white)]],
        colWidths=[3.6 * cm],
        rowHeights=[0.65 * cm],
    )
    statut_cell.setStyle(
        TableStyle(
            [
                ('BACKGROUND', (0, 0), (-1, -1), st_bg),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]
        )
    )

    qr = _qr_flowable(commande)
    right_bottom = Table([[meta_tbl], [Spacer(1, 0.2 * cm)], [statut_cell], [Spacer(1, 0.25 * cm)], [qr]])
    right_bottom.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'RIGHT')]))

    header_row = Table([[left_tbl, right_bottom]], colWidths=[W * 0.52, W * 0.48])
    header_row.setStyle(
        TableStyle(
            [
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ]
        )
    )
    story.append(header_row)
    story.append(Spacer(1, 0.35 * cm))

    band = Table(
        [[Paragraph('<font color="white"><b>BON DE COMMANDE FOURNISSEUR</b></font>', body)]],
        colWidths=[W],
        rowHeights=[0.75 * cm],
    )
    band.setStyle(
        TableStyle(
            [
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#1A56DB')),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]
        )
    )
    story.append(band)
    story.append(Spacer(1, 0.35 * cm))

    # Émetteur / Destinataire (deux colonnes comme facture)
    four_html = (
        f'<b><font color="#0F172A">{_esc_pdf(four.nom_societe)}</font></b><br/>'
        f'Réf. {_esc_pdf(four.code_fournisseur)}<br/><br/>'
        f'{_esc_pdf(four.adresse)}<br/>{_esc_pdf(four.ville)}<br/><br/>'
        f'Tél. {_esc_pdf(four.telephone)} · {_esc_pdf(four.email)}<br/>'
        f'Contact : {_esc_pdf(four.contact_nom)} · RCCM {_esc_pdf(four.rccm_id)}'
    )
    addr_tbl = Table(
        [
            [
                Paragraph('<font size="8" color="#64748B"><b>ÉMETTEUR</b></font><br/><br/>' + emitter_detail_html, body),
                Paragraph('<font size="8" color="#64748B"><b>FOURNISSEUR (DESTINATAIRE)</b></font><br/><br/>' + four_html, body),
            ]
        ],
        colWidths=[W / 2 - 0.1 * cm, W / 2 - 0.1 * cm],
    )
    addr_tbl.setStyle(
        TableStyle(
            [
                ('BOX', (0, 0), (-1, -1), 0.8, colors.HexColor('#E2E8F0')),
                ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#F8FAFC')),
                ('BACKGROUND', (1, 0), (1, 0), colors.HexColor('#F8FAFC')),
                ('LEFTPADDING', (0, 0), (-1, -1), 10),
                ('RIGHTPADDING', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]
        )
    )
    story.append(addr_tbl)
    story.append(Spacer(1, 0.45 * cm))

    dep_esc = (
        _esc_pdf(commande.depot_destination.nom)
        if commande.depot_destination_id
        else '(dépôt non renseigné)'
    )
    objet = (
        f'<i>Objet : commande de marchandises / prestations pour <b>{dep_esc}</b>'
    )
    if commande.pointdevente_destination_id:
        objet += f' — Point de vente : <b>{_esc_pdf(commande.pointdevente_destination.nom)}</b>'
    objet += '</i>'
    story.append(Paragraph(objet, muted))
    story.append(Spacer(1, 0.35 * cm))

    # Tableau lignes (style gris en-tête)
    hdr = [
        'Réf. / SKU',
        'Désignation',
        'Qté',
        'U.',
        f'PU HT ({sym})',
        'TVA %',
        f'Montant HT ({sym})',
    ]
    data_lines = [hdr]
    for ligne in commande.lignes.all():
        p = ligne.produit
        data_lines.append(
            [
                _esc_pdf((p.sku or '—')[:20]),
                _esc_pdf(p.nom[:120]),
                _esc_pdf(str(ligne.quantite_commandee)),
                _esc_pdf((ligne.unite or '')[:8]),
                _esc_pdf(f'{ligne.prix_unitaire_ht:.2f}'),
                _esc_pdf(f'{p.tva_taux:.1f}'),
                _esc_pdf(f'{ligne.sous_total_ht:.2f}'),
            ]
        )

    cw = [2.0 * cm, 6.3 * cm, 1.5 * cm, 1.0 * cm, 2.0 * cm, 1.3 * cm, 2.1 * cm]
    t_lines = Table(data_lines, colWidths=cw, repeatRows=1)
    t_lines.setStyle(
        TableStyle(
            [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E8EDF3')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 8),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#CBD5E1')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FAFBFC')]),
                ('TOPPADDING', (0, 0), (-1, -1), 6),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
                ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ]
        )
    )
    story.append(t_lines)
    story.append(Spacer(1, 0.45 * cm))

    tot_rows = [
        ['Sous-total HT', f'{commande.total_ht:.2f} {sym}'],
        ['Total TVA', f'{commande.total_tva:.2f} {sym}'],
        ['', ''],
        ['TOTAL TTC', f'{commande.total_ttc:.2f} {sym}'],
    ]
    tot_inner = Table(
        [[_esc_pdf(a), _esc_pdf(b)] for a, b in tot_rows],
        colWidths=[4.5 * cm, 4.2 * cm],
    )
    tot_inner.setStyle(
        TableStyle(
            [
                ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, -2), 'Helvetica'),
                ('FONTSIZE', (0, 0), (-1, -2), 9),
                ('LINEABOVE', (0, -1), (-1, -1), 1.2, colors.HexColor('#1A56DB')),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, -1), (-1, -1), 11),
                ('TEXTCOLOR', (0, -1), (-1, -1), colors.HexColor('#1A56DB')),
            ]
        )
    )
    tot_wrap = Table([['', tot_inner]], colWidths=[W - 8.8 * cm, 8.8 * cm])
    tot_wrap.setStyle(TableStyle([('ALIGN', (1, 0), (1, 0), 'RIGHT')]))
    story.append(tot_wrap)
    story.append(Spacer(1, 0.55 * cm))

    terms = (
        '<b>Conditions.</b> Ce document formalise la commande des articles ci-dessus aux conditions '
        'tarifaires et délais convenus avec le fournisseur. Les montants sont exprimés dans la devise '
        'indiquée ; la TVA est calculée selon le taux applicable à chaque ligne.'
    )
    story.append(Paragraph(_esc_pdf(terms), muted))
    story.append(Spacer(1, 0.35 * cm))

    if commande.notes:
        story.append(Paragraph(f'<b>Notes internes / bon :</b><br/>{_esc_pdf(commande.notes)}', body))
        story.append(Spacer(1, 0.35 * cm))

    creator = commande.cree_par
    sig_creator_img = None
    if creator:
        try:
            if getattr(creator, 'signature', None) and creator.signature:
                spath = creator.signature.path
                if os.path.isfile(spath):
                    sig_creator_img = RLImage(spath)
                    sig_creator_img.restrictSize(4.2 * cm, 2.6 * cm)
        except Exception:
            sig_creator_img = None

    left_sig_rows = [[Paragraph('<font size="8"><b>Pour ' + _esc_pdf(ent.nom) + '</b></font>', body)]]
    if creator:
        nom_cr = creator.get_full_name() or creator.get_username()
        left_sig_rows.append(
            [
                Paragraph(
                    '<font size="8">Émis par : <b>' + _esc_pdf(nom_cr) + '</b></font>',
                    body,
                )
            ]
        )
    left_sig_rows.append([Spacer(1, 0.12 * cm)])
    if sig_creator_img:
        left_sig_rows.append([sig_creator_img])
        left_sig_rows.append([Spacer(1, 0.1 * cm)])
        left_sig_rows.append(
            [
                Paragraph(
                    '<font size="8" color="#64748B">Signature et cachet</font>',
                    body,
                )
            ]
        )
    else:
        left_sig_rows.append([Spacer(1, 0.85 * cm)])
        left_sig_rows.append(
            [
                Paragraph(
                    '<font size="8" color="#64748B">Signature et cachet</font><br/>'
                    '<font size="8">_________________________</font>',
                    body,
                )
            ]
        )

    left_sig_tbl = Table(left_sig_rows, colWidths=[W / 2 - 0.15 * cm])
    left_sig_tbl.setStyle(
        TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP'), ('LEFTPADDING', (0, 0), (-1, -1), 0)])
    )

    right_sig_cell = Paragraph(
        '<font size="8"><b>Pour le fournisseur</b><br/>(' + _esc_pdf(four.nom_societe) + ')<br/><br/><br/>'
        'Signature<br/><br/>_________________________</font>',
        body,
    )
    sig = Table([[left_sig_tbl, right_sig_cell]], colWidths=[W / 2 - 0.15 * cm, W / 2 - 0.15 * cm])
    sig.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP')]))
    story.append(sig)
    story.append(Spacer(1, 0.45 * cm))

    footer_txt = getattr(settings, 'BC_PDF_FOOTER_TEXT', None)
    if footer_txt:
        story.append(Paragraph(_esc_pdf(footer_txt).replace('\n', '<br/>'), muted))
    else:
        ft = f'{ent.nom} · Document généré automatiquement — Réf. interne PK-{commande.pk}'
        story.append(Paragraph(f'<font size="7.5" color="#94A3B8">{_esc_pdf(ft)}</font>', muted))

    doc.build(story)
    buf.seek(0)
    return buf.getvalue()


def _entreprise_pour_pdf_reception(reception):
    if reception.ordre_achat_id:
        return reception.ordre_achat.entreprise
    if reception.depot_destination_id:
        return reception.depot_destination.branche.entreprise
    if reception.point_destination_id:
        return reception.point_destination.branche.entreprise
    return None


def _fournisseur_pour_pdf_reception(reception):
    if reception.ordre_achat_id:
        return reception.ordre_achat.fournisseur
    return reception.fournisseur


def _symbole_devise_reception(reception):
    if reception.ordre_achat_id and reception.ordre_achat.devise_id:
        d = reception.ordre_achat.devise
        if getattr(d, 'symbole', None):
            return d.symbole
    return '$'


def _statut_couleur_reception(statut):
    return {
        'EN_COURS': '#D97706',
        'VALIDE': '#059669',
        'ANNULE': '#DC2626',
    }.get(statut, '#64748B')


def _qr_flowable_reception(reception):
    from io import BytesIO as _BIO

    try:
        from django.conf import settings
        from django.urls import reverse

        base = getattr(settings, 'SITE_PUBLIC_URL', '').strip()
        if base:
            txt = base.rstrip('/') + reverse('achat:detail-reception', args=[reception.pk])
        else:
            ent_obj = _entreprise_pour_pdf_reception(reception)
            ent_pk = ent_obj.pk if ent_obj else 0
            txt = f'BR:{reception.numero_reception}|PK:{reception.pk}|O:{reception.ordre_achat_id or 0}|ENT:{ent_pk}'
    except Exception:
        txt = f'BR:{reception.numero_reception}|PK:{reception.pk}'

    import qrcode
    from reportlab.lib.units import cm
    from reportlab.platypus import Image as RLImage

    buf = _BIO()
    qr = qrcode.QRCode(version=None, box_size=3, border=1)
    qr.add_data(txt)
    qr.make(fit=True)
    img = qr.make_image(fill_color='#111827', back_color='white')
    img.save(buf, format='PNG')
    buf.seek(0)
    return RLImage(buf, width=2.6 * cm, height=2.6 * cm)


def build_pdf_reception(reception):
    """PDF bon de réception : même esprit que le BC (identité émetteur, fournisseur, QR, lignes détaillées, signatures)."""
    import os
    from decimal import Decimal as D

    from django.conf import settings

    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Image as RLImage
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    ent = _entreprise_pour_pdf_reception(reception)
    if not ent:
        raise ValueError("Impossible de déterminer l'entreprise pour ce bon de réception.")

    four = _fournisseur_pour_pdf_reception(reception)
    sym = _symbole_devise_reception(reception)
    bs = branche_siege_pour(ent)

    buf = BytesIO()
    margin = 1.6 * cm
    doc = SimpleDocTemplate(
        buf,
        pagesize=A4,
        leftMargin=margin,
        rightMargin=margin,
        topMargin=margin,
        bottomMargin=margin,
        title=f'BR {reception.numero_reception}',
    )
    W = A4[0] - 2 * margin
    styles = getSampleStyleSheet()
    body = ParagraphStyle(
        'body',
        parent=styles['Normal'],
        fontSize=9,
        leading=12,
        textColor=colors.HexColor('#334155'),
    )
    muted = ParagraphStyle(
        'muted',
        parent=styles['Normal'],
        fontSize=8.5,
        leading=11,
        textColor=colors.HexColor('#475569'),
    )
    title_doc = ParagraphStyle(
        'title_doc',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=11,
        textColor=colors.HexColor('#0F172A'),
        leading=14,
    )
    small_white = ParagraphStyle(
        'small_white',
        parent=styles['Normal'],
        fontName='Helvetica-Bold',
        fontSize=8,
        textColor=colors.white,
        alignment=1,
    )

    story = []

    logo_flow = None
    try:
        if ent.logo:
            pth = ent.logo.path
            if os.path.isfile(pth):
                logo_flow = RLImage(pth)
                logo_flow.restrictSize(2.5 * cm, 1.5 * cm)
    except Exception:
        logo_flow = None

    siege_txt = ''
    if bs:
        siege_txt = (
            f'<font size="9"><b>Branche siège social</b><br/>'
            f'{_esc_pdf(bs.nom)} — {_esc_pdf(bs.ville)} (code {_esc_pdf(bs.code_branche)})</font>'
        )

    emitter_detail_html = (
        f'<font size="11" color="#0F172A"><b>{_esc_pdf(ent.nom)}</b></font><br/><br/>'
        f'<font size="9">{_esc_pdf(ent.adresse_siege or "")}<br/><br/>'
        f'Tél. {_esc_pdf(ent.telephone)} · {_esc_pdf(ent.email)}<br/>'
        f'RCCM {_esc_pdf(ent.rccm)} · ID Nat. {_esc_pdf(ent.idnat)} · N° impôt {_esc_pdf(ent.numero_impot)}'
        f'</font>'
    )
    if siege_txt:
        emitter_detail_html += '<br/><br/>' + siege_txt

    emitter_compact_html = (
        f'<font size="13" color="#0F172A"><b>{_esc_pdf(ent.nom)}</b></font><br/>'
        f'<font size="8.5" color="#64748B">{_esc_pdf((ent.adresse_siege or "")[:280])}</font>'
    )

    left_stack = []
    if logo_flow:
        left_stack.append([logo_flow])
        left_stack.append([Spacer(1, 0.15 * cm)])
    left_stack.append([Paragraph(emitter_compact_html, body)])
    left_tbl = Table(left_stack, colWidths=[W * 0.52])
    left_tbl.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP'), ('LEFTPADDING', (0, 0), (-1, -1), 0)]))

    cmd = reception.ordre_achat
    meta_lines = [
        [
            Paragraph('<font size="8" color="#64748B">N° bon de réception</font>', body),
            Paragraph(f'<b>{_esc_pdf(reception.numero_reception)}</b>', title_doc),
        ],
        [
            Paragraph('<font size="8" color="#64748B">Date</font>', body),
            Paragraph(_esc_pdf(reception.date_reception.strftime('%d/%m/%Y à %H:%M')), title_doc),
        ],
        [
            Paragraph('<font size="8" color="#64748B">Bon de commande</font>', body),
            Paragraph(_esc_pdf(cmd.numero_commande if cmd else '—'), title_doc),
        ],
        [
            Paragraph('<font size="8" color="#64748B">Devise (commande)</font>', body),
            Paragraph(_esc_pdf(cmd.devise.code if cmd and cmd.devise_id else '—'), title_doc),
        ],
        [
            Paragraph('<font size="8" color="#64748B">Destination stock</font>', body),
            Paragraph(
                _esc_pdf(
                    reception.depot_destination.nom
                    if reception.depot_destination_id
                    else (
                        reception.point_destination.nom
                        if reception.point_destination_id
                        else '—'
                    )
                ),
                title_doc,
            ),
        ],
    ]

    meta_tbl = Table(meta_lines, colWidths=[3.8 * cm, 3.8 * cm])
    meta_tbl.setStyle(
        TableStyle(
            [
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('BOTTOMPADDING', (0, 0), (-1, -2), 6),
            ]
        )
    )

    st_bg = colors.HexColor(_statut_couleur_reception(reception.statut))
    statut_cell = Table(
        [[Paragraph(_esc_pdf(reception.get_statut_display()), small_white)]],
        colWidths=[3.6 * cm],
        rowHeights=[0.65 * cm],
    )
    statut_cell.setStyle(
        TableStyle(
            [
                ('BACKGROUND', (0, 0), (-1, -1), st_bg),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]
        )
    )

    qr = _qr_flowable_reception(reception)
    right_bottom = Table([[meta_tbl], [Spacer(1, 0.2 * cm)], [statut_cell], [Spacer(1, 0.25 * cm)], [qr]])
    right_bottom.setStyle(TableStyle([('ALIGN', (0, 0), (-1, -1), 'RIGHT')]))

    header_row = Table([[left_tbl, right_bottom]], colWidths=[W * 0.52, W * 0.48])
    header_row.setStyle(
        TableStyle(
            [
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 0),
                ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ]
        )
    )
    story.append(header_row)
    story.append(Spacer(1, 0.35 * cm))

    band = Table(
        [[Paragraph('<font color="white"><b>BON DE RÉCEPTION</b></font>', body)]],
        colWidths=[W],
        rowHeights=[0.75 * cm],
    )
    band.setStyle(
        TableStyle(
            [
                ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#059669')),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]
        )
    )
    story.append(band)
    story.append(Spacer(1, 0.35 * cm))

    four_html_parts = []
    if four:
        four_html_parts.extend(
            [
                f'<b><font color="#0F172A">{_esc_pdf(four.nom_societe)}</font></b><br/>',
                f'Réf. {_esc_pdf(four.code_fournisseur)}<br/><br/>',
                f'{_esc_pdf(four.adresse)}<br/>{_esc_pdf(four.ville)}<br/><br/>',
                f'Tél. {_esc_pdf(four.telephone)} · {_esc_pdf(four.email)}<br/>',
                f'Contact : {_esc_pdf(four.contact_nom)} · RCCM {_esc_pdf(four.rccm_id)}',
            ]
        )
    four_html = ''.join(four_html_parts) if four_html_parts else '<i>Pas de fournisseur renseigné</i>'
    emitter_block = Paragraph(
        '<font size="8" color="#64748B"><b>RÉCEPTIONNEUR (ENTREPRISE)</b></font><br/><br/>' + emitter_detail_html, body
    )
    addr_tbl = Table(
        [
            [
                emitter_block,
                Paragraph('<font size="8" color="#64748B"><b>FOURNISSEUR</b></font><br/><br/>' + four_html, body),
            ]
        ],
        colWidths=[W / 2 - 0.1 * cm, W / 2 - 0.1 * cm],
    )
    addr_tbl.setStyle(
        TableStyle(
            [
                ('BOX', (0, 0), (-1, -1), 0.8, colors.HexColor('#E2E8F0')),
                ('BACKGROUND', (0, 0), (0, 0), colors.HexColor('#F8FAFC')),
                ('BACKGROUND', (1, 0), (1, 0), colors.HexColor('#F8FAFC')),
                ('LEFTPADDING', (0, 0), (-1, -1), 10),
                ('RIGHTPADDING', (0, 0), (-1, -1), 10),
                ('TOPPADDING', (0, 0), (-1, -1), 10),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ]
        )
    )
    story.append(addr_tbl)
    story.append(Spacer(1, 0.4 * cm))

    trajet = ''
    if reception.depot_destination_id:
        trajet += f"Dépôt : <b>{_esc_pdf(reception.depot_destination.nom)}</b>"
        if getattr(reception.depot_destination, 'branche_id', None):
            trajet += f" ({_esc_pdf(reception.depot_destination.branche.nom)})"
    elif reception.point_destination_id:
        trajet += f"Boutique : <b>{_esc_pdf(reception.point_destination.nom)}</b>"
        if getattr(reception.point_destination, 'branche_id', None):
            trajet += f" ({_esc_pdf(reception.point_destination.branche.nom)})"
        if getattr(reception.point_destination, 'depot_source_id', None) and reception.point_destination.depot_source:
            trajet += (
                f" — Stock via dépôt source : {_esc_pdf(reception.point_destination.depot_source.nom)}"
            )
    if cmd:
        trajet += (
            '<br/><i>Aligné avec la commande : destination '
            f'{_esc_pdf(cmd.depot_destination.nom if cmd.depot_destination_id else "—")}'
        )
        if cmd.pointdevente_destination_id:
            trajet += f' / PDV {_esc_pdf(cmd.pointdevente_destination.nom)}'
        trajet += '</i>'
    story.append(
        Paragraph(
            '<font size="8.5" color="#475569">' + trajet + '</font>',
            body,
        )
    )
    story.append(Spacer(1, 0.35 * cm))

    hdr = [
        'SKU',
        'Désignation',
        'Qté reçue',
        'Écarter',
        f'PU HT ({sym})',
        f'Montant HT ({sym})',
        'TVA %',
        'Emplacement',
        'Lot',
        'Dép./Exp.',
    ]
    lignes_sorted = sorted(reception.lignes.all(), key=lambda x: x.pk)

    total_ht_lines = D('0')
    total_tva_lines = D('0')

    data_lines = [hdr]
    for ligne in lignes_sorted:
        if ligne.ligne_ordre_achat_id:
            lo = ligne.ligne_ordre_achat
            p = lo.produit
            pu = lo.prix_unitaire_ht
            qte_cmd_ref = lo.quantite_commandee
            unite_cmd = lo.unite or ''
        elif ligne.produit_id:
            lo = None
            p = ligne.produit
            pu = ligne.prix_unitaire_ht if ligne.prix_unitaire_ht is not None else p.prix_achat_ht
            qte_cmd_ref = '—'
            unite_cmd = p.unite_mesure or ''
        else:
            continue

        qrec = ligne.quantite_recue_effective or D('0')
        q_ec = ligne.quantite_ecarter or D('0')
        try:
            mt = (D(str(pu)) * qrec) if pu is not None else None
        except Exception:
            mt = None
        if mt is not None:
            total_ht_lines += mt
            tva_t = getattr(p, 'tva_taux', 0) or 0
            try:
                total_tva_lines += mt * D(str(tva_t)) / D('100')
            except Exception:
                pass

        loc_code = ligne.location.code if ligne.location_id else '—'
        lot = ligne.lot_batch or (lo.lot_batch if lo else '') or '—'
        dp = ligne.dateproduction or (lo.dateproduction if lo else None)
        de = ligne.dateexpiration or (lo.dateexpiration if lo else None)
        dpe = ''
        if dp or de:
            dpe = f'{dp.strftime("%d/%m/%Y") if dp else "—"} / {de.strftime("%d/%m/%Y") if de else "—"}'

        des = p.nom
        if ligne.marque:
            des += f' — {ligne.marque}'
        if ligne.conditionnement:
            des += f' ({ligne.conditionnement})'

        qty_cmd_txt = ''
        if lo:
            uc = (unite_cmd or '')[:8]
            qty_cmd_txt = f' (cmd {qte_cmd_ref} {uc})'

        data_lines.append(
            [
                _esc_pdf((p.sku or '—')[:18]),
                _esc_pdf((des + qty_cmd_txt)[:240]),
                _esc_pdf(str(qrec)),
                _esc_pdf(str(q_ec)),
                _esc_pdf(f'{float(pu):.2f}' if pu is not None else '—'),
                _esc_pdf(f'{float(mt):.2f}' if mt is not None else '—'),
                _esc_pdf(f'{float(p.tva_taux):.1f}' if getattr(p, 'tva_taux', None) is not None else '—'),
                _esc_pdf(loc_code[:22]),
                _esc_pdf(lot[:22]),
                _esc_pdf(dpe),
            ]
        )

    cw = [
        1.6 * cm,
        4.05 * cm,
        1.35 * cm,
        1.0 * cm,
        1.45 * cm,
        1.45 * cm,
        1.0 * cm,
        2.05 * cm,
        2.05 * cm,
        3.05 * cm,
    ]
    t_lines = Table(data_lines, colWidths=cw, repeatRows=1)
    t_lines.setStyle(
        TableStyle(
            [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E8EDF3')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#1E293B')),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 7),
                ('FONTSIZE', (0, 1), (-1, -1), 7),
                ('GRID', (0, 0), (-1, -1), 0.35, colors.HexColor('#CBD5E1')),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#FAFBFC')]),
                ('TOPPADDING', (0, 0), (-1, -1), 5),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 5),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
            ]
        )
    )
    story.append(t_lines)
    story.append(Spacer(1, 0.4 * cm))

    if total_ht_lines and total_ht_lines > 0:
        tot_rows = [
            ['Total HT (lignes)', f'{float(total_ht_lines):.2f} {sym}'],
            ['Total TVA estimée', f'{float(total_tva_lines):.2f} {sym}'],
            ['TOTAL TTC estimé', f'{float(total_ht_lines + total_tva_lines):.2f} {sym}'],
        ]
        tot_inner = Table(
            [[_esc_pdf(a), _esc_pdf(b)] for a, b in tot_rows],
            colWidths=[5.5 * cm, 3.2 * cm],
        )
        tot_inner.setStyle(
            TableStyle(
                [
                    ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
                    ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
                    ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                    ('LINEABOVE', (0, -1), (-1, -1), 1, colors.HexColor('#059669')),
                ]
            )
        )
        tot_wrap = Table([['', tot_inner]], colWidths=[W - 9.0 * cm, 9.0 * cm])
        tot_wrap.setStyle(TableStyle([('ALIGN', (1, 0), (1, 0), 'RIGHT')]))
        story.append(tot_wrap)
        story.append(Spacer(1, 0.35 * cm))

    story.append(
        Paragraph(
            _esc_pdf(
                'Les quantités réceptionnées seront mouvementées en stock après validation '
                'du bon. Contrôlez écarts et lots avant signature.'
            ),
            muted,
        )
    )

    if reception.notes:
        story.append(Spacer(1, 0.3 * cm))
        story.append(Paragraph(f'<b>Notes bon :</b><br/>{_esc_pdf(reception.notes)}', body))

    story.append(Spacer(1, 0.5 * cm))

    def _sig_block(user_field, titre):
        if not user_field:
            return Paragraph(
                f'<font size="8">{_esc_pdf(titre)}</font><br/><br/>________________________________',
                body,
            )
        img = None
        try:
            if getattr(user_field, 'signature', None) and user_field.signature:
                spath = user_field.signature.path
                if os.path.isfile(spath):
                    img = RLImage(spath)
                    img.restrictSize(4 * cm, 2.5 * cm)
        except Exception:
            img = None
        nom = user_field.get_full_name() or user_field.get_username()
        rows_sig = [[Paragraph(f'<font size="8"><b>{_esc_pdf(titre)}</b></font><br/><font size="8">{_esc_pdf(nom)}</font>', body)]]
        rows_sig.append([Spacer(1, 0.1 * cm)])
        if img:
            rows_sig.append([img])
        else:
            rows_sig.append([Spacer(1, 1 * cm)])
        rows_sig.append(
            [
                Paragraph(
                    '<font size="8" color="#64748B">Signature</font>',
                    body,
                )
            ]
        )
        return Table(rows_sig, colWidths=[W / 2 - 0.15 * cm])

    left_sb = _sig_block(reception.cree_par, 'Créé par')
    right_sb = _sig_block(reception.recu_par, 'Réceptionné par')
    sig_tbl = Table([[left_sb, right_sb]], colWidths=[W / 2 - 0.15 * cm, W / 2 - 0.15 * cm])
    sig_tbl.setStyle(TableStyle([('VALIGN', (0, 0), (-1, -1), 'TOP')]))
    story.append(sig_tbl)
    story.append(Spacer(1, 0.45 * cm))

    footer_txt = getattr(settings, 'BR_PDF_FOOTER_TEXT', None) or getattr(
        settings, 'BC_PDF_FOOTER_TEXT', None
    )
    if footer_txt:
        story.append(Paragraph(_esc_pdf(footer_txt).replace('\n', '<br/>'), muted))
    else:
        ft = f'{ent.nom} — Bon de réception — Réf. {reception.numero_reception} (PK-{reception.pk})'
        story.append(Paragraph(f'<font size="7.5" color="#94A3B8">{_esc_pdf(ft)}</font>', muted))

    doc.build(story)
    buf.seek(0)
    return buf.getvalue()
