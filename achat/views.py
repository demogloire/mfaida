import io
import json
from decimal import Decimal, InvalidOperation
from datetime import date, datetime

import openpyxl
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.db import transaction
from django.db.models import Prefetch, Q
from django.http import HttpResponse
from django.utils.dateparse import parse_date
from django.urls import reverse
from django.views.decorators.http import require_POST
from django_htmx.http import HttpResponseClientRedirect
from openpyxl.styles import Alignment, Font, PatternFill

from utilisateur.decorators import login_requis
from entreprise.models import Entreprise, Depot, PointVente, Devise, Produit, Location
from tiers.models import Fournisseur

from .bc_export import build_excel_bc, build_pdf_bc, build_pdf_reception
from stock.reception_stock import appliquer_bon_reception_au_stock

from .models import OrdreAchat, LigneOrdreAchat, BonReception, LigneBonReception
from .forms import (
    OrdreAchatForm,
    LigneOrdreAchatForm,
    BonReceptionForm,
    ImportLignesBcForm,
    LigneBonReceptionFormSet,
)


# ───────────────────────────── helpers ─────────────────────────────

def _get_entreprise(user):
    """
    Entreprise liée au compte (même logique que la gestion utilisateurs / rôles).
    Branche d'affectation, sinon entreprise dont l'utilisateur est propriétaire (Entreprise.user).
    Pas de repli sur « la première entreprise » pour éviter les fuites de périmètre.
    """
    if getattr(user, 'branche_id', None):
        return user.branche.entreprise
    return Entreprise.objects.filter(user=user).first()


def _est_admin(user):
    return getattr(user, 'admin', False) or getattr(user, 'is_superuser', False)


def _pdv_ids_par_depot_json(entreprise, admin, user=None):
    """
    Pour chaque dépôt, liste des PK de points de vente actifs dont depot_source = ce dépôt
    (ordre alphabétique sur le nom du PDV — le premier est présélectionné côté client).
    Non-admins : seuls les dépôts et PDV visibles via AccesDepot / AccesPointVente (peut_voir).
    """
    from entreprise.models import Branche, PointVente

    if admin:
        branches = Branche.objects.filter(est_actif=True)
    elif entreprise and user is not None and getattr(user, 'is_authenticated', False):
        from stock.access import queryset_depots_visibles, queryset_points_vente_visibles

        depots = queryset_depots_visibles(user, entreprise, False)
        pvs = queryset_points_vente_visibles(user, entreprise, False).filter(
            depot_source__isnull=False,
        ).order_by('depot_source_id', 'nom')
        d = {}
        depot_ids_allowed = set(depots.values_list('pk', flat=True))
        for pv in pvs:
            ds = pv.depot_source_id
            if ds is not None and ds in depot_ids_allowed:
                d.setdefault(ds, []).append(pv.pk)
        return json.dumps({str(k): v for k, v in d.items()})
    elif entreprise:
        branches = entreprise.branches.filter(est_actif=True)
    else:
        return json.dumps({})

    d = {}
    qs = PointVente.objects.filter(
        branche__in=branches,
        est_actif=True,
        depot_source__isnull=False,
    ).order_by('depot_source_id', 'nom')
    for pv in qs:
        d.setdefault(pv.depot_source_id, []).append(pv.pk)
    return json.dumps({str(k): v for k, v in d.items()})


def _exiger_entreprise_si_besoin(request, admin, entreprise, redirect_name='achat:liste-commandes'):
    """Pour un utilisateur non admin : entreprise obligatoire pour agir sur les commandes."""
    if admin or entreprise:
        return None
    messages.error(request, "Aucune entreprise associée à votre compte. Impossible d'utiliser les achats.")
    return redirect(redirect_name)


def _parse_decimal_bc(val, default=None):
    """Parse un nombre depuis Excel (float, int, str avec virgule ou point)."""
    if val is None or val == '':
        return default
    try:
        if isinstance(val, bool):
            return default
        if isinstance(val, (int, float)):
            return Decimal(str(val))
        return Decimal(str(val).replace(',', '.').strip().replace(' ', '').replace('\xa0', ''))
    except (InvalidOperation, ValueError, TypeError):
        return default


def _normalize_excel_cell_text(val):
    """
    Texte / SKU / code depuis Excel : évite '2.0' au lieu de '2', préserve les chaînes non numériques.
    """
    if val is None or val == '':
        return ''
    if isinstance(val, bool):
        return '1' if val else '0'
    if isinstance(val, float):
        if val != val:  # NaN
            return ''
        if val == int(val):
            return str(int(val))
        return str(val).strip()
    if isinstance(val, int):
        return str(val)
    s = str(val).strip()
    # Chaîne type "2.0" issue d'Excel
    if len(s) >= 3 and s.endswith('.0') and s[:-2].lstrip('-').isdigit():
        return s[:-2]
    return s


def _sku_lookup_entreprise(ent_id, sku_token):
    """Trouve un produit par SKU avec tolérance zéros à gauche (Excel numérique vs SKU texte)."""
    if not sku_token:
        return None
    tokens = {sku_token}
    if sku_token.isdigit():
        for pad in range(2, 8):
            tokens.add(sku_token.zfill(pad))
    qs = Produit.objects.filter(entreprise_id=ent_id, est_actif=True)
    for t in tokens:
        p = qs.filter(sku=t).first()
        if p:
            return p
    return None


def _parse_date_bc(val):
    if val is None or val == '':
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    if not s:
        return None
    try:
        return datetime.strptime(s[:10], '%Y-%m-%d').date()
    except ValueError:
        pass
    for fmt in ('%d/%m/%Y', '%d-%m-%Y'):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except ValueError:
            continue
    return None


def _resolve_location_code(commande, code_raw):
    from entreprise.models import Location

    code = str(code_raw or '').strip()
    if not code:
        return None
    if commande.depot_destination_id:
        return Location.objects.filter(
            branche_id=commande.depot_destination.branche_id,
            code__iexact=code,
        ).first()
    branches = commande.entreprise.branches.all()
    return Location.objects.filter(branche__in=branches, code__iexact=code).first()


def _resolve_produit_bc(ent_id, data):
    cb = _normalize_excel_cell_text(data.get('code_barre'))
    sku_raw = _normalize_excel_cell_text(data.get('sku'))
    nv = data.get('nom')
    if nv is None or nv == '':
        nom = ''
    elif isinstance(nv, str):
        nom = nv.strip()
    else:
        nom = str(nv).strip()

    if not cb and not sku_raw and not nom:
        return None, 'missing'
    if cb:
        p = Produit.objects.select_related('entreprise').filter(code_barre=cb).first()
        if not p:
            return None, 'introuvable_cb'
        if p.entreprise_id != ent_id:
            return None, 'autre_entreprise_cb'
        if not p.est_actif:
            return None, 'inactif'
        return p, None
    if sku_raw:
        p = _sku_lookup_entreprise(ent_id, sku_raw)
        if p:
            return p, None
        # Excel a peut‑être dégradé le SKU (nombre) ou nom plus fiable
        if nom:
            p = Produit.objects.filter(
                entreprise_id=ent_id, nom__iexact=nom, est_actif=True
            ).first()
            if p:
                return p, None
        return None, 'introuvable_sku'
    p = Produit.objects.filter(entreprise_id=ent_id, nom__iexact=nom, est_actif=True).first()
    if not p:
        return None, 'introuvable_nom'
    return p, None


def _msg_erreur_produit(code):
    return {
        'introuvable_cb': 'code-barres inconnu.',
        'autre_entreprise_cb': 'ce code-barres appartient à un autre catalogue.',
        'introuvable_sku': (
            'SKU introuvable pour cette entreprise '
            '(Excel transforme souvent les SKU en nombre : formater la colonne en « Texte », ou vérifier le nom exact).'
        ),
        'introuvable_nom': 'nom de produit introuvable (exact, insensible à la casse).',
        'inactif': 'produit inactif.',
        'missing': 'renseigner au moins code_barre, sku ou nom.',
    }.get(code, code)


def _row_import_vide(row):
    if row is None:
        return True
    for v in row:
        if v is None:
            continue
        if isinstance(v, str):
            if v.strip():
                return False
            continue
        if v != '':
            return False
    return True


def _commande_export_base_qs():
    return OrdreAchat.objects.select_related(
        'entreprise',
        'fournisseur',
        'depot_destination__branche',
        'pointdevente_destination__branche',
        'devise',
        'cree_par',
    ).prefetch_related(
        Prefetch(
            'lignes',
            queryset=LigneOrdreAchat.objects.select_related(
                'produit',
                'location',
            ).order_by('pk'),
        )
    )


# ═══════════════════════════════════════════════════════════════════
#  ORDRES D'ACHAT
# ═══════════════════════════════════════════════════════════════════

@login_requis
def liste_commandes(request):
    entreprise = _get_entreprise(request.user)
    admin = _est_admin(request.user)
    if admin:
        qs = OrdreAchat.objects.all().select_related(
            'fournisseur',
            'entreprise',
            'depot_destination',
            'pointdevente_destination',
            'devise',
        )
    elif entreprise:
        qs = OrdreAchat.objects.filter(entreprise=entreprise).select_related(
            'fournisseur',
            'entreprise',
            'depot_destination',
            'pointdevente_destination',
            'devise',
        )
    else:
        qs = OrdreAchat.objects.none()

    q = request.GET.get('q', '')
    statut = request.GET.get('statut', '')
    if q:
        qs = qs.filter(
            Q(numero_commande__icontains=q)
            | Q(fournisseur__nom_societe__icontains=q)
            | Q(entreprise__nom__icontains=q)
            | Q(depot_destination__nom__icontains=q)
            | Q(pointdevente_destination__nom__icontains=q)
        )
    if statut:
        qs = qs.filter(statut=statut)

    ctx = {
        'commandes': qs,
        'actif': 'commandes',
        'STATUTS': OrdreAchat.STATUT_CHOICES,
        'q': q, 'statut': statut,
        'admin': admin,
        'entreprise_bc_scope': entreprise,
    }
    if request.htmx and request.htmx.target == 'table-container':
        return render(request, 'achat/commandes/partial/table.html', ctx)
    return render(request, 'achat/commandes/liste.html', ctx)


@login_requis
def detail_commande(request, pk):
    entreprise = _get_entreprise(request.user)
    admin = _est_admin(request.user)
    if admin:
        commande = get_object_or_404(OrdreAchat, pk=pk)
    else:
        redir = _exiger_entreprise_si_besoin(request, admin, entreprise)
        if redir:
            return redir
        commande = get_object_or_404(OrdreAchat, pk=pk, entreprise=entreprise)
    lignes = commande.lignes.select_related('produit', 'location')
    ctx = {
        'commande': commande,
        'lignes': lignes,
        'actif': 'commandes',
        'peut_modifier': commande.statut == 'BROUILLON',
        'import_lignes_form': ImportLignesBcForm() if commande.statut == 'BROUILLON' else None,
        'admin': admin,
        'entreprise_bc_scope': entreprise,
    }
    return render(request, 'achat/commandes/detail.html', ctx)


@login_requis
def export_commande_excel(request, pk):
    entreprise = _get_entreprise(request.user)
    admin = _est_admin(request.user)
    qs = _commande_export_base_qs()
    if admin:
        commande = get_object_or_404(qs, pk=pk)
    else:
        redir = _exiger_entreprise_si_besoin(request, admin, entreprise)
        if redir:
            return redir
        commande = get_object_or_404(qs, pk=pk, entreprise=entreprise)

    data = build_excel_bc(commande)
    fn = f"BC_{commande.numero_commande.replace('/', '-')}_donnees.xlsx"
    response = HttpResponse(
        data,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{fn}"'
    return response


@login_requis
def export_commande_pdf(request, pk):
    entreprise = _get_entreprise(request.user)
    admin = _est_admin(request.user)
    qs = _commande_export_base_qs()
    if admin:
        commande = get_object_or_404(qs, pk=pk)
    else:
        redir = _exiger_entreprise_si_besoin(request, admin, entreprise)
        if redir:
            return redir
        commande = get_object_or_404(qs, pk=pk, entreprise=entreprise)

    data = build_pdf_bc(commande)
    fn = f"BC_{commande.numero_commande.replace('/', '-')}.pdf"
    response = HttpResponse(data, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{fn}"'
    return response


@login_requis
def creer_commande(request):
    entreprise = _get_entreprise(request.user)
    admin = _est_admin(request.user)
    redir = _exiger_entreprise_si_besoin(request, admin, entreprise)
    if redir:
        return redir
    form = OrdreAchatForm(request.POST or None, entreprise=entreprise, admin=admin, user=request.user)
    if request.method == 'POST' and form.is_valid():
        commande = form.save(commit=False)
        # Admin : l'entreprise est déduite du fournisseur choisi
        if admin:
            commande.entreprise = commande.fournisseur.entreprise
        else:
            commande.entreprise = entreprise
        commande.cree_par = request.user
        commande.save()
        messages.success(request, f"Bon de commande {commande.numero_commande} créé.")
        return redirect('achat:detail-commande', pk=commande.pk)
    ctx = {
        'form': form,
        'actif': 'commandes',
        'titre': 'Nouveau bon de commande',
        'entreprise_affichage': None if admin else entreprise,
        'pdv_par_depot_json': _pdv_ids_par_depot_json(entreprise, admin, request.user),
    }
    return render(request, 'achat/commandes/form.html', ctx)


@login_requis
def modifier_commande(request, pk):
    entreprise = _get_entreprise(request.user)
    admin = _est_admin(request.user)
    if admin:
        commande = get_object_or_404(OrdreAchat, pk=pk, statut='BROUILLON')
    else:
        redir = _exiger_entreprise_si_besoin(request, admin, entreprise)
        if redir:
            return redir
        commande = get_object_or_404(OrdreAchat, pk=pk, entreprise=entreprise, statut='BROUILLON')
    form = OrdreAchatForm(request.POST or None, instance=commande, entreprise=entreprise, admin=admin, user=request.user)
    if request.method == 'POST' and form.is_valid():
        form.save()
        messages.success(request, "Commande mise à jour.")
        return redirect('achat:detail-commande', pk=pk)
    ctx = {
        'form': form,
        'actif': 'commandes',
        'titre': 'Modifier la commande',
        'commande': commande,
        'entreprise_affichage': None if admin else entreprise,
        'pdv_par_depot_json': _pdv_ids_par_depot_json(entreprise, admin, request.user),
    }
    return render(request, 'achat/commandes/form.html', ctx)


@login_requis
@require_POST
def envoyer_commande(request, pk):
    """Passe le BC de Brouillon → Envoyé au fournisseur (pas d’envoi e-mail automatique : utiliser export PDF/mail hors appli)."""
    entreprise = _get_entreprise(request.user)
    admin = _est_admin(request.user)
    if admin:
        commande = get_object_or_404(OrdreAchat, pk=pk, statut='BROUILLON')
    else:
        redir = _exiger_entreprise_si_besoin(request, admin, entreprise)
        if redir:
            return redir
        commande = get_object_or_404(OrdreAchat, pk=pk, entreprise=entreprise, statut='BROUILLON')
    if not commande.lignes.exists():
        messages.error(request, 'Ajoutez au moins une ligne de produit avant d’envoyer le bon.')
        return redirect('achat:detail-commande', pk=pk)
    commande.statut = 'ENVOYE'
    commande.save(update_fields=['statut'])
    messages.success(
        request,
        'Bon marqué « Envoyé au fournisseur ». '
        'Téléchargez le PDF ou l’Excel pour le transmettre au fournisseur (e-mail, plateforme, etc.). '
        'Vous pouvez maintenant créer une réception liée à cette commande.',
    )
    return redirect('achat:detail-commande', pk=pk)


@login_requis
def annuler_commande(request, pk):
    entreprise = _get_entreprise(request.user)
    admin = _est_admin(request.user)
    if admin:
        commande = get_object_or_404(OrdreAchat, pk=pk)
    else:
        redir = _exiger_entreprise_si_besoin(request, admin, entreprise)
        if redir:
            return redir
        commande = get_object_or_404(OrdreAchat, pk=pk, entreprise=entreprise)
    if request.method == 'POST' and commande.statut in ('BROUILLON', 'ENVOYE'):
        commande.statut = 'ANNULE'
        commande.save(update_fields=['statut'])
        messages.warning(request, f"Commande {commande.numero_commande} annulée.")
        return redirect('achat:liste-commandes')
    ctx = {'commande': commande}
    return render(request, 'achat/commandes/confirm_annuler.html', ctx)


# ── Lignes (HTMX inline) ──

@login_requis
def ajouter_ligne(request, pk):
    entreprise = _get_entreprise(request.user)
    admin = _est_admin(request.user)
    cmd_base = OrdreAchat.objects.select_related(
        'depot_destination__branche__entreprise',
        'entreprise',
    )
    if admin:
        commande = get_object_or_404(cmd_base, pk=pk, statut='BROUILLON')
    else:
        redir = _exiger_entreprise_si_besoin(request, admin, entreprise)
        if redir:
            return redir
        commande = get_object_or_404(
            cmd_base, pk=pk, entreprise=entreprise, statut='BROUILLON'
        )
    form = LigneOrdreAchatForm(request.POST or None, commande=commande)
    if request.method == 'POST' and form.is_valid():
        cleaned = form.cleaned_data
        _, mode = LigneOrdreAchat.creer_ou_fusionner(commande, cleaned)
        if mode == 'merged':
            messages.info(
                request,
                'Quantité ajoutée sur la ligne existante (même produit et mêmes informations).',
            )
        commande.recalculer_totaux()
        lignes = commande.lignes.select_related('produit')
        return render(request, 'achat/commandes/partial/lignes.html', {'commande': commande, 'lignes': lignes})
    return render(request, 'achat/commandes/partial/form_ligne.html', {'form': form, 'commande': commande})


@login_requis
def supprimer_ligne(request, pk):
    ligne = get_object_or_404(LigneOrdreAchat.objects.select_related('ordre_achat'), pk=pk)
    commande = ligne.ordre_achat
    entreprise = _get_entreprise(request.user)
    admin = _est_admin(request.user)
    if not admin:
        if not entreprise or commande.entreprise_id != entreprise.pk:
            messages.error(request, "Vous ne pouvez pas modifier cette commande.")
            lignes = commande.lignes.select_related('produit')
            return render(request, 'achat/commandes/partial/lignes.html', {'commande': commande, 'lignes': lignes})

    lignes = commande.lignes.select_related('produit')
    if commande.statut != 'BROUILLON':
        messages.warning(request, "Seuls les bons en brouillon peuvent être modifiés.")
        return render(request, 'achat/commandes/partial/lignes.html', {'commande': commande, 'lignes': lignes})

    qte_recue = ligne.quantite_recue if ligne.quantite_recue is not None else Decimal('0')
    if qte_recue != Decimal('0'):
        messages.warning(
            request,
            "Impossible de supprimer cette ligne : une quantité a déjà été réceptionnée.",
        )
        return render(request, 'achat/commandes/partial/lignes.html', {'commande': commande, 'lignes': lignes})

    ligne.delete()
    commande.recalculer_totaux()
    lignes = commande.lignes.select_related('produit')
    messages.success(request, "Ligne supprimée.")
    return render(request, 'achat/commandes/partial/lignes.html', {'commande': commande, 'lignes': lignes})


@login_requis
def import_lignes_commande(request, pk):
    """Import Excel des lignes sur un BC brouillon (catalogue = entreprise du bon)."""
    entreprise = _get_entreprise(request.user)
    admin = _est_admin(request.user)
    cmd_base = OrdreAchat.objects.select_related(
        'depot_destination__branche__entreprise',
        'entreprise',
    )
    if admin:
        commande = get_object_or_404(cmd_base, pk=pk, statut='BROUILLON')
    else:
        redir = _exiger_entreprise_si_besoin(request, admin, entreprise)
        if redir:
            return redir
        commande = get_object_or_404(
            cmd_base, pk=pk, entreprise=entreprise, statut='BROUILLON'
        )

    if request.method != 'POST':
        return redirect('achat:detail-commande', pk=pk)

    form = ImportLignesBcForm(request.POST, request.FILES)
    if not form.is_valid():
        for err in form.errors.get('fichier', []):
            messages.error(request, err)
        return redirect('achat:detail-commande', pk=pk)

    ent_id = commande.entreprise_id
    fichier = form.cleaned_data['fichier']
    merged = created = skipped = 0
    errors = []

    try:
        wb = openpyxl.load_workbook(fichier, data_only=True)
        ws = wb.active
        headers = [
            str(c.value).strip().lower().replace(' ', '_') if c.value else ''
            for c in ws[1]
        ]

        with transaction.atomic():
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                data = dict(zip(headers, row))
                if _row_import_vide(row):
                    skipped += 1
                    continue

                qte = _parse_decimal_bc(data.get('quantite_commandee'))
                prix = _parse_decimal_bc(data.get('prix_unitaire_ht'))
                if qte is None or qte <= 0:
                    errors.append(f"Ligne {row_idx}: quantité commandée invalide ou manquante.")
                    continue
                if prix is None or prix < 0:
                    errors.append(f"Ligne {row_idx}: prix unitaire HT invalide ou manquant.")
                    continue

                produit, err_code = _resolve_produit_bc(ent_id, data)
                if not produit:
                    hint = _msg_erreur_produit(err_code)
                    errors.append(f"Ligne {row_idx}: {hint}")
                    continue

                loc = _resolve_location_code(commande, data.get('location_code'))
                cleaned = {
                    'produit': produit,
                    'quantite_commandee': qte,
                    'prix_unitaire_ht': prix,
                    'unite': str(data.get('unite') or '').strip(),
                    'lot_batch': str(data.get('lot_batch') or '').strip(),
                    'dateproduction': _parse_date_bc(data.get('dateproduction')),
                    'dateexpiration': _parse_date_bc(data.get('dateexpiration')),
                    'location': loc,
                }

                if cleaned['location'] is None and str(data.get('location_code') or '').strip():
                    errors.append(
                        f"Ligne {row_idx}: emplacement « {data.get('location_code')} » introuvable pour ce dépôt / cette entreprise."
                    )
                    continue

                _, mode = LigneOrdreAchat.creer_ou_fusionner(commande, cleaned)
                if mode == 'merged':
                    merged += 1
                else:
                    created += 1

            commande.recalculer_totaux()

    except Exception as e:
        messages.error(request, f"Lecture du fichier impossible : {e}")
        return redirect('achat:detail-commande', pk=pk)

    if created or merged:
        messages.success(
            request,
            f"{created} ligne(s) créée(s), {merged} fusion(s) avec une ligne existante.",
        )
    if errors:
        messages.warning(
            request,
            f"{len(errors)} ligne(s) en erreur — détails : « {' ; '.join(errors[:8])}{'…' if len(errors) > 8 else ''} »",
        )
    if skipped and not (created or merged) and not errors:
        messages.info(request, "Aucune ligne importée (lignes vides ignorées).")

    return redirect('achat:detail-commande', pk=pk)


@login_requis
def telecharger_modele_lignes_bc(request, pk):
    entreprise = _get_entreprise(request.user)
    admin = _est_admin(request.user)
    if admin:
        commande = get_object_or_404(OrdreAchat, pk=pk, statut='BROUILLON')
    else:
        redir = _exiger_entreprise_si_besoin(request, admin, entreprise)
        if redir:
            return redir
        commande = get_object_or_404(OrdreAchat, pk=pk, entreprise=entreprise, statut='BROUILLON')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Lignes_BC'

    # Pas de code-barres : SKU = code seul (référence unique par entreprise), pas un libellé type SKU-Nom-UM.
    headers = [
        'sku',
        'nom',
        'quantite_commandee',
        'prix_unitaire_ht',
        'unite',
        'lot_batch',
        'dateproduction',
        'dateexpiration',
        'location_code',
    ]
    h_font = Font(bold=True, color='FFFFFF')
    h_fill = PatternFill('solid', fgColor='1A56DB')
    h_align = Alignment(horizontal='center', vertical='center')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = h_font
        cell.fill = h_fill
        cell.alignment = h_align
        ws.column_dimensions[cell.column_letter].width = 22

    lignes_list = list(
        commande.lignes.select_related('produit', 'location').order_by('pk')
    )
    if lignes_list:
        for ligne in lignes_list:
            p = ligne.produit
            sku_val = (p.sku or '').strip()
            ws.append([
                sku_val,
                p.nom,
                ligne.quantite_commandee,
                ligne.prix_unitaire_ht,
                ligne.unite or '',
                ligne.lot_batch or '',
                ligne.dateproduction,
                ligne.dateexpiration,
                ligne.location.code if ligne.location_id else '',
            ])
    else:
        ws.append([
            'MON-SKU-001',
            'Exemple produit',
            10,
            2.5,
            'PCS',
            '',
            '',
            '',
            '',
        ])
        ws.append([
            'REF-002',
            'Autre exemple',
            5,
            100,
            'KG',
            'LOT-A',
            '',
            '',
            '',
        ])

    ws2 = wb.create_sheet('Instructions')
    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 72
    ws2.append(['Colonne', 'Obligatoire', 'Description'])
    ws2['A1'].font = Font(bold=True)
    ws2['B1'].font = Font(bold=True)
    ws2['C1'].font = Font(bold=True)
    for row in [
        (
            'sku',
            'NON',
            "Code SKU uniquement (comme en catalogue), sans nom ni unité — unique par entreprise. "
            "Si renseigné : identifie le produit en priorité.",
        ),
        (
            'nom',
            'NON',
            "Si SKU vide : nom exact du produit (catalogue de l'entreprise du BC). Au moins sku ou nom doit identifier le produit.",
        ),
        ('quantite_commandee', 'OUI', 'Quantité à commander (nombre > 0).'),
        ('prix_unitaire_ht', 'OUI', 'Prix unitaire hors taxes.'),
        ('unite', 'NON', 'Unité de commande (ex. PCS, KG) — défaut vide.'),
        ('lot_batch', 'NON', 'Lot / batch — défaut vide.'),
        ('dateproduction', 'NON', 'Date AAAA-MM-JJ ou JJ/MM/AAAA — optionnel.'),
        ('dateexpiration', 'NON', 'Idem — optionnel.'),
        ('location_code', 'NON', "Code d'emplacement (Location) du dépôt / de l'entreprise — optionnel."),
        ('—', '—', "Colonne « code_barres » absente : l'import peut encore accepter une colonne code_barre pour compatibilité."),
    ]:
        ws2.append(row)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    safe_num = commande.numero_commande.replace('/', '-')
    suffix = 'lignes' if lignes_list else 'modele'
    response['Content-Disposition'] = (
        f'attachment; filename="{suffix}_lignes_bc_{safe_num}.xlsx"'
    )
    return response


# ═══════════════════════════════════════════════════════════════════
#  RECEPTIONS
# ═══════════════════════════════════════════════════════════════════

def _reception_scope_q(entreprise):
    return (
        Q(ordre_achat__entreprise=entreprise)
        | Q(depot_destination__branche__entreprise=entreprise)
        | Q(point_destination__branche__entreprise=entreprise)
    )


def _branche_id_destination(dep, pv):
    if dep:
        return dep.branche_id
    if pv:
        return pv.branche_id
    return None


def _locations_pour_branche(branche_id):
    from entreprise.models import Location

    if branche_id is None:
        return Location.objects.none()
    return Location.objects.filter(branche_id=branche_id).order_by('code')


def _locations_pour_bon(reception):
    bid = _branche_id_destination(reception.depot_destination, reception.point_destination)
    return _locations_pour_branche(bid)


def _locations_catalogue_entreprise(entreprise, admin=False):
    """Toutes les locations des branches actives d'une entreprise (filtrage UI par branche côté client)."""
    from entreprise.models import Location

    if not entreprise:
        return Location.objects.none()
    return Location.objects.filter(
        branche__entreprise=entreprise,
        branche__est_actif=True,
    ).select_related('branche').order_by('branche__nom', 'code')


def _locations_branches_du_bc(commande):
    """Emplacements limités aux branches du dépôt et/ou du PDV portés par le bon de commande."""
    from entreprise.models import Location

    branche_ids = []
    if commande.depot_destination_id:
        branche_ids.append(commande.depot_destination.branche_id)
    if commande.pointdevente_destination_id:
        branche_ids.append(commande.pointdevente_destination.branche_id)
    if not branche_ids:
        return _locations_catalogue_entreprise(commande.entreprise)
    return Location.objects.filter(
        branche_id__in=set(branche_ids),
        branche__est_actif=True,
    ).select_related('branche').order_by('branche__nom', 'code')


def _resolver_location_ligne(pk_raw, branche_id):
    from entreprise.models import Location

    if branche_id is None or pk_raw is None or str(pk_raw).strip() == '':
        return None
    try:
        lid = int(pk_raw)
    except (ValueError, TypeError):
        return None
    return Location.objects.filter(pk=lid, branche_id=branche_id).first()


def _problemes_validation_reception(reception):
    problèmes = []
    lignes = list(reception.lignes.all())
    if not any((l.quantite_recue_effective or Decimal('0')) > 0 for l in lignes):
        problèmes.append(
            'Le bon doit comporter au moins une ligne avec une quantité reçue positive.'
        )
    for l in lignes:
        q = l.quantite_recue_effective or Decimal('0')
        if q > 0 and not l.location_id:
            if l.produit_id:
                pnom = l.produit.nom
            elif l.ligne_ordre_achat_id:
                pnom = l.ligne_ordre_achat.produit.nom
            else:
                pnom = 'une ligne'
            problèmes.append(f'Emplacement manquant pour : {pnom}')
    return problèmes


def _decimal_post_br(val, default=None):
    if val is None or str(val).strip() == '':
        return default
    try:
        return Decimal(str(val).replace(',', '.').strip())
    except (InvalidOperation, ValueError, TypeError):
        return default


def _ligne_indices_reception_simple():
    return list(range(20))


def _ligne_rows_reception_simple_from_post(request, ligne_indices):
    rows = []
    for idx in ligne_indices:
        rows.append({
            'idx': idx,
            'produit': (request.POST.get(f'produit_{idx}', '') or '').strip(),
            'qte': request.POST.get(f'qte_{idx}', '') or '',
            'prix': request.POST.get(f'prix_{idx}', '') or '',
            'ecart': request.POST.get(f'ecart_{idx}', '') or '0',
            'lot': request.POST.get(f'lot_{idx}', '') or '',
            'marque': request.POST.get(f'marque_{idx}', '') or '',
            'cond': request.POST.get(f'cond_{idx}', '') or '',
            'dprod': request.POST.get(f'dprod_{idx}', '') or '',
            'dexp': request.POST.get(f'dexp_{idx}', '') or '',
            'loc': (request.POST.get(f'loc_{idx}', '') or '').strip(),
        })
    return rows


def _ligne_rows_reception_simple_blank(ligne_indices):
    return [{
        'idx': idx,
        'produit': '',
        'qte': '',
        'prix': '',
        'ecart': '0',
        'lot': '',
        'marque': '',
        'cond': '',
        'dprod': '',
        'dexp': '',
        'loc': '',
    } for idx in ligne_indices]


def _render_reception_simple_panel(
    request,
    *,
    entreprise,
    form,
    produits_qs,
    ligne_rows,
    admin=False,
    succes_reception=None,
    locations_catalogue=None,
):
    """Fragment HTMX ou inclusion page complète : formulaire réception simple."""
    locs = (
        locations_catalogue
        if locations_catalogue is not None
        else _locations_catalogue_entreprise(entreprise)
    )
    return render(
        request,
        'achat/receptions/partial/reception_simple_panel.html',
        {
            'form': form,
            'produits': produits_qs,
            'ligne_rows': ligne_rows,
            'entreprise': entreprise,
            'locations_catalogue': locs,
            'succes_reception': succes_reception,
        },
    )


def _ligne_indices_ajout_details_reception():
    return list(range(15))


def _entreprise_depuis_reception(reception):
    if reception.ordre_achat_id:
        return reception.ordre_achat.entreprise
    if reception.depot_destination_id:
        return reception.depot_destination.branche.entreprise
    if reception.point_destination_id:
        return reception.point_destination.branche.entreprise
    return None


def _synchroniser_statut_commande_quantites(commande):
    if commande.statut not in ('ENVOYE', 'RECU_PARTIEL', 'RECU_TOTAL'):
        return
    lignes = list(commande.lignes.all())
    if not lignes:
        return
    completes = sum(
        1 for l in lignes if (l.quantite_recue or Decimal('0')) >= l.quantite_commandee
    )
    if completes == len(lignes):
        nouveau = 'RECU_TOTAL'
    elif any((l.quantite_recue or Decimal('0')) > 0 for l in lignes):
        nouveau = 'RECU_PARTIEL'
    else:
        nouveau = 'ENVOYE'
    if commande.statut != nouveau:
        commande.statut = nouveau
        commande.save(update_fields=['statut'])


def _decimal_restant_ligne_ordre_bc(lo):
    """Quantité encore ouverte sur une ligne du bon (= commandée − cumul réceptionné sur tous les BR)."""
    r = lo.quantite_commandee - (lo.quantite_recue or Decimal('0'))
    return max(Decimal('0'), r)


def _lignes_commande_et_restants(lignes_list):
    """Préparation affichage formulaire création réception depuis BC."""
    return [
        {'ligne': ligne, 'restant_bc': _decimal_restant_ligne_ordre_bc(ligne)}
        for ligne in lignes_list
    ]


def _charger_reception_pour_detail(request, pk):
    """Retourne (reception, None) ou (None, HttpResponseRedirect) si accès refusé."""
    entreprise_user = _get_entreprise(request.user)
    admin = _est_admin(request.user)
    base = BonReception.objects.select_related(
        'ordre_achat__fournisseur',
        'ordre_achat__entreprise',
        'fournisseur',
        'depot_destination__branche__entreprise',
        'point_destination__branche__entreprise',
        'recu_par',
        'cree_par',
    )
    if admin:
        return get_object_or_404(base, pk=pk), None
    redir = _exiger_entreprise_si_besoin(
        request, admin, entreprise_user, redirect_name='achat:liste-receptions'
    )
    if redir:
        return None, redir
    reception = get_object_or_404(
        base.filter(_reception_scope_q(entreprise_user)),
        pk=pk,
    )
    return reception, None


def _context_lignes_reception_detail(request, reception):
    ent_br = _entreprise_depuis_reception(reception)
    lignes_qs = list(
        reception.lignes.select_related(
            'ligne_ordre_achat__produit',
            'produit',
            'location',
        )
    )
    for lbr in lignes_qs:
        lbr.restant_bc_ligne = None
        lbr.qte_max_modal = None
        if lbr.ligne_ordre_achat_id:
            lc = lbr.ligne_ordre_achat
            old_here = lbr.quantite_recue_effective or Decimal('0')
            rs = _decimal_restant_ligne_ordre_bc(lc)
            lbr.restant_bc_ligne = rs
            lbr.qte_max_modal = rs + old_here

    lignes_bc_ajout_rows = []
    if reception.ordre_achat_id:
        lignes_bc_ajout_rows = [
            {'lc': lc, 'restant': _decimal_restant_ligne_ordre_bc(lc)}
            for lc in reception.ordre_achat.lignes.select_related('produit').order_by('pk')
        ]

    produits_ajout = (
        Produit.objects.filter(entreprise=ent_br, est_actif=True).order_by('nom')
        if ent_br
        else Produit.objects.none()
    )

    return {
        'reception': reception,
        'lignes': lignes_qs,
        'ligne_add_indices': _ligne_indices_ajout_details_reception(),
        'produits_ajout': produits_ajout,
        'lignes_bc_ajout_rows': lignes_bc_ajout_rows,
        'locations_bon': _locations_pour_bon(reception),
        'actif': 'receptions',
    }


def _render_fragment_lignes_reception(request, reception):
    return render(
        request,
        'achat/receptions/partial/reception_lignes_et_ajout.html',
        _context_lignes_reception_detail(request, reception),
    )


@login_requis
def liste_receptions(request):
    entreprise = _get_entreprise(request.user)
    admin = _est_admin(request.user)
    rel = (
        'ordre_achat__fournisseur', 'ordre_achat__entreprise', 'fournisseur',
        'depot_destination', 'point_destination', 'recu_par', 'cree_par',
    )
    if admin:
        qs = BonReception.objects.all().select_related(*rel)
    elif entreprise:
        qs = BonReception.objects.filter(_reception_scope_q(entreprise)).select_related(*rel)
    else:
        qs = BonReception.objects.none()

    q = request.GET.get('q', '')
    statut = request.GET.get('statut', '')
    if q:
        qs = qs.filter(
            Q(numero_reception__icontains=q)
            | Q(ordre_achat__numero_commande__icontains=q)
            | Q(ordre_achat__fournisseur__nom_societe__icontains=q)
            | Q(fournisseur__nom_societe__icontains=q)
        )
    if statut:
        qs = qs.filter(statut=statut)

    ctx = {
        'receptions': qs,
        'actif': 'receptions',
        'STATUTS': BonReception.STATUT_CHOICES,
        'q': q, 'statut': statut,
    }
    if request.htmx and request.htmx.target == 'table-container':
        return render(request, 'achat/receptions/partial/table.html', ctx)
    return render(request, 'achat/receptions/liste.html', ctx)


@login_requis
def export_reception_pdf(request, pk):
    reception, redir = _charger_reception_pour_detail(request, pk)
    if redir:
        return redir
    reception = (
        BonReception.objects.select_related(
            'ordre_achat__entreprise',
            'ordre_achat__fournisseur',
            'ordre_achat__devise',
            'ordre_achat__depot_destination',
            'ordre_achat__pointdevente_destination',
            'fournisseur',
            'depot_destination__branche',
            'point_destination__branche',
            'point_destination__depot_source',
            'cree_par',
            'recu_par',
        )
        .prefetch_related(
            Prefetch(
                'lignes',
                queryset=LigneBonReception.objects.select_related(
                    'produit',
                    'ligne_ordre_achat__produit',
                    'ligne_ordre_achat__ordre_achat',
                    'location',
                ).order_by('pk'),
            )
        )
        .get(pk=reception.pk)
    )
    data = build_pdf_reception(reception)
    safe_num = reception.numero_reception.replace('/', '-')
    fn = f'BR_{safe_num}.pdf'
    response = HttpResponse(data, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="{fn}"'
    return response


@login_requis
@require_POST
def supprimer_reception(request, pk):
    reception, redir = _charger_reception_pour_detail(request, pk)
    if redir:
        return redir
    if reception.statut == 'VALIDE':
        messages.error(request, 'Impossible de supprimer une réception déjà validée (stock mouvementé).')
        return redirect('achat:liste-receptions')
    cmd = reception.ordre_achat
    cmd_pk_for_sync = cmd.pk if cmd else None

    try:
        with transaction.atomic():
            br = BonReception.objects.select_for_update().get(pk=reception.pk)
            if br.statut == 'VALIDE':
                raise ValueError('Ce bon est validé : suppression impossible.')
            ordre_lié = br.ordre_achat_id
            for lb in (
                LigneBonReception.objects.select_for_update()
                .filter(bon_reception_id=br.pk)
                .select_related('ligne_ordre_achat')
            ):
                if lb.ligne_ordre_achat_id:
                    lo = LigneOrdreAchat.objects.select_for_update().get(pk=lb.ligne_ordre_achat_id)
                    if ordre_lié and lo.ordre_achat_id != ordre_lié:
                        raise ValueError('Incohérence bon de réception / ligne de commande.')
                    q = lb.quantite_recue_effective or Decimal('0')
                    lo.quantite_recue = max(Decimal('0'), (lo.quantite_recue or Decimal('0')) - q)
                    lo.save(update_fields=['quantite_recue'])
            br.delete()
            if cmd_pk_for_sync:
                commande = OrdreAchat.objects.get(pk=cmd_pk_for_sync)
                _synchroniser_statut_commande_quantites(commande)
    except ValueError as exc:
        messages.error(request, str(exc))
        return redirect('achat:detail-reception', pk=pk)

    messages.success(request, 'Bon de réception supprimé.')
    return redirect('achat:liste-receptions')


@login_requis
def creer_reception(request, ordre_pk):
    entreprise = _get_entreprise(request.user)
    admin = _est_admin(request.user)
    base_q = OrdreAchat.objects.select_related(
        'fournisseur', 'entreprise', 'depot_destination', 'pointdevente_destination',
    )
    commande_candidate = base_q.filter(pk=ordre_pk).first()
    if commande_candidate is None:
        messages.error(request, 'Bon de commande introuvable.')
        return redirect('achat:liste-commandes')

    if not admin:
        redir = _exiger_entreprise_si_besoin(request, admin, entreprise, redirect_name='achat:liste-receptions')
        if redir:
            return redir
        if commande_candidate.entreprise_id != entreprise.pk:
            messages.error(request, 'Ce bon de commande ne fait pas partie de votre entreprise.')
            return redirect('achat:liste-commandes')
    else:
        # Admin : voit tous les BC ; création de réception limitée au même périmètre qu’un utilisateur rattaché.
        if not entreprise:
            messages.error(
                request,
                'Pour créer une réception, votre compte doit être rattaché à une entreprise (branche ou fiche entreprise).',
            )
            return redirect('achat:liste-commandes')
        if commande_candidate.entreprise_id != entreprise.pk:
            messages.error(request, 'La réception est limitée aux bons de commande de votre entreprise.')
            return redirect('achat:liste-commandes')

    if commande_candidate.statut not in ('ENVOYE', 'RECU_PARTIEL'):
        messages.warning(
            request,
            'La réception n’est disponible que pour une commande « Envoyée » ou « Partiellement reçue ». '
            'En brouillon, validez ou envoyez d’abord le bon depuis sa fiche.',
        )
        return redirect('achat:detail-commande', pk=ordre_pk)

    commande = commande_candidate
    if not commande.depot_destination_id and not commande.pointdevente_destination_id:
        messages.error(
            request,
            'Le bon de commande n’a pas de dépôt ni de point de vente de destination : complétez la commande avant de réceptionner.',
        )
        return redirect('achat:detail-commande', pk=ordre_pk)

    lignes_commande = list(commande.lignes.select_related('produit'))
    if not lignes_commande:
        messages.error(request, 'Ce bon de commande n’a aucune ligne : impossible de réceptionner.')
        return redirect('achat:detail-commande', pk=ordre_pk)

    lignes_bc_creer = _lignes_commande_et_restants(lignes_commande)

    locations_lignes_bc = lambda: _locations_branches_du_bc(commande)

    if request.method == 'POST':
        form = BonReceptionForm(
            request.POST,
            commande=commande,
            entreprise=entreprise,
            admin=admin,
            user=request.user,
        )
        if form.is_valid():
            cleaned = form.cleaned_data
            branche_id = _branche_id_destination(
                cleaned.get('depot_destination'),
                cleaned.get('point_destination'),
            )
            planned = []
            for ligne in lignes_commande:
                raw_qte = request.POST.get(f'qte_{ligne.pk}')
                if raw_qte is None or str(raw_qte).strip() == '':
                    continue
                qte = _decimal_post_br(raw_qte, Decimal('0')) or Decimal('0')
                if qte <= 0:
                    continue
                loc = _resolver_location_ligne(request.POST.get(f'loc_{ligne.pk}'), branche_id)
                if not loc:
                    messages.error(
                        request,
                        f'Choisissez un emplacement valide pour « {ligne.produit.nom} » '
                        '(même branche que la destination de réception).',
                    )
                    return render(request, 'achat/receptions/form.html', {
                        'form': form,
                        'commande': commande,
                        'lignes_bc_creer': _lignes_commande_et_restants(
                            list(commande.lignes.select_related('produit')),
                        ),
                        'locations_lignes': locations_lignes_bc(),
                        'actif': 'receptions',
                    })
                ecarter = _decimal_post_br(request.POST.get(f'ecart_{ligne.pk}', 0), Decimal('0')) or Decimal('0')
                if ecarter > qte:
                    messages.error(
                        request,
                        f'Pour « {ligne.produit.nom} » : la quantité à l’écart ne peut pas dépasser la quantité reçue.',
                    )
                    return render(request, 'achat/receptions/form.html', {
                        'form': form,
                        'commande': commande,
                        'lignes_bc_creer': _lignes_commande_et_restants(
                            list(commande.lignes.select_related('produit')),
                        ),
                        'locations_lignes': locations_lignes_bc(),
                        'actif': 'receptions',
                    })
                dp = parse_date(request.POST.get(f'dprod_{ligne.pk}', '') or '')
                de = parse_date(request.POST.get(f'dexp_{ligne.pk}', '') or '')
                planned.append((
                    ligne,
                    qte,
                    ecarter,
                    (request.POST.get(f'lot_{ligne.pk}', '') or '')[:20],
                    (request.POST.get(f'marque_{ligne.pk}', '') or '')[:100],
                    (request.POST.get(f'cond_{ligne.pk}', '') or '')[:100],
                    dp,
                    de,
                    loc,
                ))

            if not planned:
                messages.error(
                    request,
                    'Ajoutez au moins une ligne avec une quantité reçue positive et un emplacement.',
                )
                return render(request, 'achat/receptions/form.html', {
                    'form': form,
                    'commande': commande,
                    'lignes_bc_creer': _lignes_commande_et_restants(
                        list(commande.lignes.select_related('produit')),
                    ),
                    'locations_lignes': locations_lignes_bc(),
                    'actif': 'receptions',
                })

            try:
                with transaction.atomic():
                    cmd_lock = OrdreAchat.objects.select_for_update().get(pk=commande.pk)

                    for (
                        ligne,
                        qte,
                        ecarter,
                        lot_batch,
                        marque,
                        conditionnement,
                        dp,
                        de,
                        loc,
                    ) in planned:
                        lo = LigneOrdreAchat.objects.select_for_update().get(
                            pk=ligne.pk,
                            ordre_achat_id=cmd_lock.pk,
                        )
                        restant_lc = _decimal_restant_ligne_ordre_bc(lo)
                        if qte > restant_lc:
                            raise ValueError(
                                f'Pour « {lo.produit.nom} » : quantité ({qte}) supérieure au restant disponible '
                                f'sur cette ligne du bon ({restant_lc}).'
                            )

                    reception = form.save(commit=False)
                    reception.ordre_achat = cmd_lock
                    reception.fournisseur = None
                    reception.recu_par = request.user
                    reception.cree_par = request.user
                    reception.save()

                    for (
                        ligne,
                        qte,
                        ecarter,
                        lot_batch,
                        marque,
                        conditionnement,
                        dp,
                        de,
                        loc,
                    ) in planned:
                        LigneBonReception.objects.create(
                            bon_reception=reception,
                            ligne_ordre_achat=ligne,
                            quantite_recue_effective=qte,
                            quantite_ecarter=ecarter,
                            motif_ecarter='',
                            lot_batch=lot_batch,
                            marque=marque,
                            conditionnement=conditionnement,
                            dateproduction=dp,
                            dateexpiration=de,
                            location=loc,
                        )
                        lo = LigneOrdreAchat.objects.select_for_update().get(pk=ligne.pk)
                        lo.quantite_recue = (lo.quantite_recue or Decimal('0')) + qte
                        lo.save(update_fields=['quantite_recue'])

                    _synchroniser_statut_commande_quantites(cmd_lock)

            except ValueError as exc:
                messages.error(request, str(exc))
                return render(request, 'achat/receptions/form.html', {
                    'form': form,
                    'commande': commande,
                    'lignes_bc_creer': _lignes_commande_et_restants(
                        list(commande.lignes.select_related('produit')),
                    ),
                    'locations_lignes': locations_lignes_bc(),
                    'actif': 'receptions',
                })

            messages.success(request, f"Bon de réception {reception.numero_reception} créé.")
            return redirect('achat:detail-reception', pk=reception.pk)
    else:
        form = BonReceptionForm(
            commande=commande,
            entreprise=entreprise,
            admin=admin,
            user=request.user,
        )

    ctx = {
        'form': form,
        'commande': commande,
        'lignes_bc_creer': lignes_bc_creer,
        'locations_lignes': locations_lignes_bc(),
        'actif': 'receptions',
    }
    return render(request, 'achat/receptions/form.html', ctx)


@login_requis
def creer_reception_simple(request):
    """Réception hors bon de commande — périmètre entreprise utilisateur ; enregistrement HTMX sans rechargement complet."""
    ligne_indices = _ligne_indices_reception_simple()
    entreprise = _get_entreprise(request.user)
    admin_flag = _est_admin(request.user)

    if not entreprise and not admin_flag:
        messages.error(
            request,
            "Une entreprise doit être associée à votre compte (branche ou fiche entreprise) pour créer une réception simple.",
        )
        return redirect('achat:liste-receptions')

    if admin_flag and not entreprise:
        produits_qs = Produit.objects.filter(est_actif=True).select_related(
            'sous_categorie__categorie__entreprise',
        ).order_by('entreprise_id', 'nom')

        def locations_qs_global():
            return Location.objects.filter(
                branche__est_actif=True,
            ).select_related('branche').order_by('branche__nom', 'code')
    else:
        produits_qs = Produit.objects.filter(entreprise=entreprise, est_actif=True).order_by('nom')

        def locations_qs_global():
            return _locations_catalogue_entreprise(entreprise)

    form_scope_admin = admin_flag and not entreprise

    if request.method == 'POST':
        form = BonReceptionForm(
            request.POST,
            commande=None,
            entreprise=entreprise,
            admin=form_scope_admin,
            user=request.user,
        )
        ligne_rows = _ligne_rows_reception_simple_from_post(request, ligne_indices)

        def _erreur_fragment():
            locs = locations_qs_global()
            if request.htmx:
                return _render_reception_simple_panel(
                    request,
                    entreprise=entreprise,
                    form=form,
                    produits_qs=produits_qs,
                    ligne_rows=ligne_rows,
                    admin=admin_flag,
                    locations_catalogue=locs,
                )
            return render(request, 'achat/receptions/form_simple.html', {
                'form': form,
                'produits': produits_qs,
                'ligne_rows': ligne_rows,
                'entreprise': entreprise,
                'locations_catalogue': locs,
                'succes_reception': None,
                'actif': 'receptions',
            })

        if not form.is_valid():
            messages.error(request, 'Corrigez les erreurs ci-dessous.')
            return _erreur_fragment()

        cleaned_fd = form.cleaned_data
        branche_id = _branche_id_destination(
            cleaned_fd.get('depot_destination'),
            cleaned_fd.get('point_destination'),
        )

        lignes_ok = []
        erreur_parse = None
        for idx in ligne_indices:
            pid = request.POST.get(f'produit_{idx}', '').strip()
            if not pid:
                continue
            try:
                pid_int = int(pid)
            except ValueError:
                continue
            pq = Produit.objects.filter(pk=pid_int, est_actif=True)
            if entreprise:
                pq = pq.filter(entreprise=entreprise)
            produit = pq.first()
            if not produit:
                erreur_parse = f'Produit #{pid_int} invalide ou inaccessible.'
                break
            qte = _decimal_post_br(request.POST.get(f'qte_{idx}', ''), None)
            if qte is None or qte <= 0:
                continue
            loc = _resolver_location_ligne(request.POST.get(f'loc_{idx}'), branche_id)
            if not loc:
                erreur_parse = (
                    f'Pour « {produit.nom} », choisissez un emplacement de la même branche '
                    'que la destination (dépôt ou boutique).'
                )
                break
            ecarter = _decimal_post_br(request.POST.get(f'ecart_{idx}', 0), Decimal('0')) or Decimal('0')
            if ecarter > qte:
                erreur_parse = (
                    f'Pour « {produit.nom} », la quantité à l’écart ne peut pas dépasser la quantité reçue.'
                )
                break
            lignes_ok.append({
                'produit': produit,
                'qte': qte,
                'ecarter': ecarter,
                'lot': (request.POST.get(f'lot_{idx}', '') or '')[:20],
                'marque': (request.POST.get(f'marque_{idx}', '') or '')[:100],
                'cond': (request.POST.get(f'cond_{idx}', '') or '')[:100],
                'dprod': parse_date(request.POST.get(f'dprod_{idx}', '') or ''),
                'dexp': parse_date(request.POST.get(f'dexp_{idx}', '') or ''),
                'prix': _decimal_post_br(request.POST.get(f'prix_{idx}', ''), None),
                'location': loc,
            })

        if erreur_parse:
            messages.error(request, erreur_parse)
            return _erreur_fragment()

        if not lignes_ok:
            messages.error(request, 'Ajoutez au moins une ligne avec quantité et emplacement.')
            return _erreur_fragment()

        with transaction.atomic():
            reception = form.save(commit=False)
            reception.ordre_achat = None
            reception.recu_par = request.user
            reception.cree_par = request.user
            reception.save()
            numero = reception.numero_reception
            pk_out = reception.pk
            for row in lignes_ok:
                px = row['prix']
                if px is not None and px < 0:
                    px = None
                LigneBonReception.objects.create(
                    bon_reception=reception,
                    ligne_ordre_achat=None,
                    produit=row['produit'],
                    prix_unitaire_ht=px,
                    quantite_recue_effective=row['qte'],
                    quantite_ecarter=row['ecarter'],
                    motif_ecarter='',
                    lot_batch=row['lot'],
                    marque=row['marque'],
                    conditionnement=row['cond'],
                    dateproduction=row['dprod'],
                    dateexpiration=row['dexp'],
                    location=row['location'],
                )

        messages.success(
            request,
            f'Bon de réception {numero} créé. Vous pouvez poursuivre la saisie des produits depuis la fiche du bon.',
        )
        detail_url = reverse('achat:detail-reception', args=[pk_out])
        if request.htmx:
            return HttpResponseClientRedirect(detail_url)
        return redirect(detail_url)

    form = BonReceptionForm(
        commande=None,
        entreprise=entreprise,
        admin=form_scope_admin,
        user=request.user,
    )
    ligne_rows = _ligne_rows_reception_simple_blank(ligne_indices)
    locs = locations_qs_global()
    return render(request, 'achat/receptions/form_simple.html', {
        'form': form,
        'produits': produits_qs,
        'ligne_rows': ligne_rows,
        'entreprise': entreprise,
        'locations_catalogue': locs,
        'succes_reception': None,
        'actif': 'receptions',
    })


@login_requis
def detail_reception(request, pk):
    reception, redir = _charger_reception_pour_detail(request, pk)
    if redir:
        return redir
    ctx = _context_lignes_reception_detail(request, reception)
    return render(request, 'achat/receptions/detail.html', ctx)


@login_requis
@require_POST
def ajouter_lignes_reception(request, pk):
    reception, redir = _charger_reception_pour_detail(request, pk)
    if redir:
        return redir

    if reception.statut != 'EN_COURS':
        messages.error(request, 'Seules les réceptions « en cours » acceptent de nouvelles lignes.')
        if request.htmx:
            return _render_fragment_lignes_reception(request, reception)
        return redirect('achat:detail-reception', pk=pk)

    ent_br = _entreprise_depuis_reception(reception)
    if not ent_br:
        messages.error(request, 'Impossible de déterminer l’entreprise de ce bon.')
        if request.htmx:
            return _render_fragment_lignes_reception(request, reception)
        return redirect('achat:detail-reception', pk=pk)

    ligne_add_indices = _ligne_indices_ajout_details_reception()
    crees = 0

    try:
        with transaction.atomic():
            recep = BonReception.objects.select_for_update().get(pk=reception.pk)
            if recep.statut != 'EN_COURS':
                raise ValueError('Statut modifié entre-temps.')
            recep.recu_par_id = request.user.pk
            recep.save(update_fields=['recu_par'])
            bid_br = _branche_id_destination(recep.depot_destination, recep.point_destination)
            if recep.ordre_achat_id:
                commande = OrdreAchat.objects.select_for_update().get(pk=recep.ordre_achat_id)
                if commande.entreprise_id != ent_br.pk:
                    raise ValueError('Incohérence entreprise / commande.')
                for lc in commande.lignes.select_related('produit'):
                    qte_key = f'qte_{lc.pk}'
                    raw_qte = request.POST.get(qte_key)
                    if raw_qte is None or str(raw_qte).strip() == '':
                        continue
                    qte = _decimal_post_br(raw_qte, Decimal('0')) or Decimal('0')
                    if qte <= 0:
                        continue
                    lo = LigneOrdreAchat.objects.select_for_update().get(pk=lc.pk, ordre_achat_id=commande.pk)
                    restant = _decimal_restant_ligne_ordre_bc(lo)
                    if qte > restant:
                        raise ValueError(
                            f'Pour « {lo.produit.nom} » : quantité ({qte}) supérieure au restant ({restant}) '
                            f'sur cette ligne du bon de commande.'
                        )
                    loc = _resolver_location_ligne(request.POST.get(f'loc_{lc.pk}'), bid_br)
                    if not loc:
                        raise ValueError(
                            f'Emplacement invalide ou manquant pour « {lo.produit.nom} ».'
                        )
                    ecarter = _decimal_post_br(request.POST.get(f'ecart_{lc.pk}', 0), Decimal('0')) or Decimal('0')
                    if ecarter > qte:
                        raise ValueError(
                            f'Pour « {lo.produit.nom} » : la quantité à l’écart ne peut pas dépasser la quantité reçue.'
                        )
                    dp = parse_date(request.POST.get(f'dprod_{lc.pk}', '') or '')
                    de = parse_date(request.POST.get(f'dexp_{lc.pk}', '') or '')
                    LigneBonReception.objects.create(
                        bon_reception=recep,
                        ligne_ordre_achat=lo,
                        quantite_recue_effective=qte,
                        quantite_ecarter=ecarter,
                        motif_ecarter='',
                        lot_batch=(request.POST.get(f'lot_{lc.pk}', '') or '')[:20],
                        marque=(request.POST.get(f'marque_{lc.pk}', '') or '')[:100],
                        conditionnement=(request.POST.get(f'cond_{lc.pk}', '') or '')[:100],
                        dateproduction=dp,
                        dateexpiration=de,
                        location=loc,
                    )
                    lo.quantite_recue = (lo.quantite_recue or Decimal('0')) + qte
                    lo.save(update_fields=['quantite_recue'])
                    crees += 1
                _synchroniser_statut_commande_quantites(commande)
            else:
                for idx in ligne_add_indices:
                    pid = request.POST.get(f'nproduit_{idx}', '').strip()
                    if not pid:
                        continue
                    try:
                        pid_int = int(pid)
                    except ValueError:
                        continue
                    produit = Produit.objects.filter(pk=pid_int, entreprise=ent_br, est_actif=True).first()
                    if not produit:
                        raise ValueError(f'Produit #{pid_int} invalide.')
                    qte = _decimal_post_br(request.POST.get(f'nqte_{idx}', ''), None)
                    if qte is None or qte <= 0:
                        continue
                    loc = _resolver_location_ligne(request.POST.get(f'nloc_{idx}'), bid_br)
                    if not loc:
                        raise ValueError(
                            f'Emplacement invalide ou manquant pour « {produit.nom} ».'
                        )
                    ecarter = _decimal_post_br(request.POST.get(f'necart_{idx}', 0), Decimal('0')) or Decimal('0')
                    if ecarter > qte:
                        raise ValueError(
                            f'Pour « {produit.nom} » : la quantité à l’écart ne peut pas dépasser la quantité reçue.'
                        )
                    dp = parse_date(request.POST.get(f'ndprod_{idx}', '') or '')
                    de = parse_date(request.POST.get(f'ndexp_{idx}', '') or '')
                    px = _decimal_post_br(request.POST.get(f'nprix_{idx}', ''), None)
                    if px is not None and px < 0:
                        px = None
                    LigneBonReception.objects.create(
                        bon_reception=recep,
                        ligne_ordre_achat=None,
                        produit=produit,
                        prix_unitaire_ht=px,
                        quantite_recue_effective=qte,
                        quantite_ecarter=ecarter,
                        motif_ecarter='',
                        lot_batch=(request.POST.get(f'nlot_{idx}', '') or '')[:20],
                        marque=(request.POST.get(f'nmarque_{idx}', '') or '')[:100],
                        conditionnement=(request.POST.get(f'ncond_{idx}', '') or '')[:100],
                        dateproduction=dp,
                        dateexpiration=de,
                        location=loc,
                    )
                    crees += 1
    except ValueError as exc:
        messages.error(request, str(exc))
        if request.htmx:
            reception.refresh_from_db()
            return _render_fragment_lignes_reception(request, reception)
        return redirect('achat:detail-reception', pk=pk)

    if crees == 0:
        messages.warning(
            request,
            'Aucune ligne créée : indiquez des quantités reçues et un emplacement par ligne.',
        )
    else:
        messages.success(request, f'{crees} ligne(s) ajoutée(s) au bon {reception.numero_reception}.')

    reception.refresh_from_db()
    if request.htmx:
        return _render_fragment_lignes_reception(request, reception)
    return redirect('achat:detail-reception', pk=pk)



@login_requis
@require_POST
def modifier_ligne_reception(request, pk):
    reception, redir = _charger_reception_pour_detail(request, pk)
    if redir:
        return redir
    if reception.statut != 'EN_COURS':
        messages.error(request, 'Modification impossible pour ce bon.')
        return redirect('achat:detail-reception', pk=pk)
    ligne_pk = request.POST.get('ligne_pk')
    ligne = get_object_or_404(LigneBonReception, pk=ligne_pk, bon_reception=reception)
    bid = _branche_id_destination(reception.depot_destination, reception.point_destination)
    loc = _resolver_location_ligne(request.POST.get('location'), bid)
    if not loc:
        messages.error(request, 'Choisissez un emplacement valide pour cette ligne.')
        if request.htmx:
            return _render_fragment_lignes_reception(request, reception)
        return redirect('achat:detail-reception', pk=pk)
    new_qte = _decimal_post_br(request.POST.get('quantite_recue_effective'), None)
    if new_qte is None or new_qte <= 0:
        messages.error(request, 'Indiquez une quantité reçue positive.')
        if request.htmx:
            return _render_fragment_lignes_reception(request, reception)
        return redirect('achat:detail-reception', pk=pk)
    ecarter = _decimal_post_br(request.POST.get('quantite_ecarter', 0), Decimal('0')) or Decimal('0')
    if ecarter > new_qte:
        messages.error(request, "La quantité à l'écart ne peut pas dépasser la quantité reçue.")
        if request.htmx:
            return _render_fragment_lignes_reception(request, reception)
        return redirect('achat:detail-reception', pk=pk)
    dp = parse_date(request.POST.get('dateproduction', '') or '')
    de = parse_date(request.POST.get('dateexpiration', '') or '')

    try:
        with transaction.atomic():
            br = BonReception.objects.select_for_update().get(pk=reception.pk)
            if br.statut != 'EN_COURS':
                raise ValueError('Statut modifié entre-temps.')
            br.recu_par_id = request.user.pk
            br.save(update_fields=['recu_par'])
            lb = LigneBonReception.objects.select_for_update().get(pk=ligne.pk, bon_reception_id=br.pk)
            old_q = lb.quantite_recue_effective or Decimal('0')
            delta = new_qte - old_q

            if lb.ligne_ordre_achat_id:
                lo = LigneOrdreAchat.objects.select_for_update().get(pk=lb.ligne_ordre_achat_id)
                restant_bc = _decimal_restant_ligne_ordre_bc(lo)
                max_sur_br = restant_bc + old_q
                if new_qte > max_sur_br:
                    raise ValueError(
                        f'Pour « {lo.produit.nom} » : vous ne pouvez pas saisir plus de '
                        f'{max_sur_br} sur cette ligne (reste encore à couvrir sur le bon de commande : {restant_bc}).'
                    )
                cand = (lo.quantite_recue or Decimal('0')) + delta
                if cand < 0:
                    raise ValueError('Quantité incompatible avec le cumul du bon de commande.')
                lo.quantite_recue = cand
                lo.save(update_fields=['quantite_recue'])

            lb.quantite_recue_effective = new_qte
            lb.quantite_ecarter = ecarter
            lb.motif_ecarter = ''
            lb.lot_batch = (request.POST.get('lot_batch', '') or '')[:20]
            lb.marque = (request.POST.get('marque', '') or '')[:100]
            lb.conditionnement = (request.POST.get('conditionnement', '') or '')[:100]
            lb.dateproduction = dp
            lb.dateexpiration = de
            lb.location = loc
            lb.save()
            if lb.ligne_ordre_achat_id:
                _synchroniser_statut_commande_quantites(lo.ordre_achat)
    except ValueError as exc:
        messages.error(request, str(exc))
        if request.htmx:
            reception.refresh_from_db()
            return _render_fragment_lignes_reception(request, reception)
        return redirect('achat:detail-reception', pk=pk)

    messages.success(request, 'Ligne mise à jour.')
    reception.refresh_from_db()
    if request.htmx:
        return _render_fragment_lignes_reception(request, reception)
    return redirect('achat:detail-reception', pk=pk)


@login_requis
@require_POST
def supprimer_ligne_reception(request, pk):
    reception, redir = _charger_reception_pour_detail(request, pk)
    if redir:
        return redir
    if reception.statut != 'EN_COURS':
        messages.error(request, 'Suppression impossible pour ce bon.')
        return redirect('achat:detail-reception', pk=pk)
    ligne_pk = request.POST.get('ligne_pk')
    ligne = get_object_or_404(LigneBonReception, pk=ligne_pk, bon_reception=reception)

    try:
        with transaction.atomic():
            br = BonReception.objects.select_for_update().get(pk=reception.pk)
            if br.statut != 'EN_COURS':
                raise ValueError('Statut modifié entre-temps.')
            br.recu_par_id = request.user.pk
            br.save(update_fields=['recu_par'])
            lb = LigneBonReception.objects.select_for_update().get(pk=ligne.pk, bon_reception_id=br.pk)
            q = lb.quantite_recue_effective or Decimal('0')
            cmd = None
            if lb.ligne_ordre_achat_id:
                lo = LigneOrdreAchat.objects.select_for_update().get(pk=lb.ligne_ordre_achat_id)
                cmd = lo.ordre_achat
                lo.quantite_recue = max(Decimal('0'), (lo.quantite_recue or Decimal('0')) - q)
                lo.save(update_fields=['quantite_recue'])
            lb.delete()
            if cmd:
                _synchroniser_statut_commande_quantites(cmd)
    except ValueError as exc:
        messages.error(request, str(exc))
        if request.htmx:
            reception.refresh_from_db()
            return _render_fragment_lignes_reception(request, reception)
        return redirect('achat:detail-reception', pk=pk)

    messages.success(request, 'Ligne retirée du bon.')
    reception.refresh_from_db()
    if request.htmx:
        return _render_fragment_lignes_reception(request, reception)
    return redirect('achat:detail-reception', pk=pk)


@login_requis
def valider_reception(request, pk):
    entreprise = _get_entreprise(request.user)
    admin = _est_admin(request.user)

    qs_base = BonReception.objects.select_related(
        'ordre_achat__fournisseur',
        'fournisseur',
        'depot_destination',
        'point_destination',
        'recu_par',
        'cree_par',
    ).prefetch_related(
        'lignes__produit',
        'lignes__ligne_ordre_achat__produit',
    )
    if admin:
        reception = get_object_or_404(qs_base, pk=pk, statut='EN_COURS')
    else:
        redir = _exiger_entreprise_si_besoin(request, admin, entreprise, redirect_name='achat:liste-receptions')
        if redir:
            return redir
        reception = get_object_or_404(
            qs_base.filter(_reception_scope_q(entreprise)),
            pk=pk,
            statut='EN_COURS',
        )
    if request.method == 'POST':
        try:
            with transaction.atomic():
                recep = BonReception.objects.select_for_update().get(pk=reception.pk)
                if recep.statut != 'EN_COURS':
                    raise ValueError("Ce bon n'est plus en cours.")
                audit_fields = []
                if not recep.recu_par_id:
                    recep.recu_par = request.user
                    audit_fields.append('recu_par')
                if not recep.cree_par_id:
                    recep.cree_par = request.user
                    audit_fields.append('cree_par')
                if audit_fields:
                    recep.save(update_fields=audit_fields)
                probs = _problemes_validation_reception(recep)
                if probs:
                    raise ValueError(' ; '.join(probs))
                appliquer_bon_reception_au_stock(recep, recep.recu_par or request.user)
                recep.statut = 'VALIDE'
                recep.save(update_fields=['statut'])
        except ValueError as exc:
            msg = str(exc)
            if "n'est plus en cours" in msg:
                messages.warning(request, msg)
            else:
                messages.error(request, msg)
            return redirect('achat:detail-reception', pk=pk)
        try:
            from datetime import date

            from finance.posting import montant_piece_reception, poster_entree_stock_ohada

            mt = montant_piece_reception(recep)
            ent_br = (
                recep.depot_destination.branche.entreprise
                if recep.depot_destination_id
                else recep.point_destination.branche.entreprise
            )
            poster_entree_stock_ohada(
                ent_br,
                mt,
                recep.numero_reception,
                date.today(),
                f'Réception {recep.numero_reception}',
                request.user,
            )
        except Exception as exc:
            messages.warning(request, f'Comptabilité OHADA simplifiée : {exc}')
        messages.success(
            request,
            f"Réception validée ({recep.numero_reception}) — les mouvements de stock sont enregistrés.",
        )
        return redirect('achat:detail-reception', pk=pk)
    problemes_validation = _problemes_validation_reception(reception)
    ctx = {
        'reception': reception,
        'bloque_validation': bool(problemes_validation),
        'problemes_validation': problemes_validation,
    }
    return render(request, 'achat/receptions/confirm_valider.html', ctx)
