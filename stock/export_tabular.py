"""Exports tableaux génériques (Excel / PDF) pour les rapports stock."""
from __future__ import annotations

import io
import re
from decimal import Decimal
from typing import Any, Iterable, List, Sequence, Tuple

from django.http import HttpResponse


def fichier_nom_safe_fragment(text: str, max_len: int = 60) -> str:
    s = re.sub(r'[^\w\-]+', '_', (text or '').strip())[:max_len].strip('_')
    return s or 'export'


def cell_str(val: Any) -> str:
    if val is None:
        return ''
    if isinstance(val, Decimal):
        return format(val, 'f')
    if isinstance(val, bool):
        return 'Oui' if val else 'Non'
    s = str(val).strip()
    return re.sub(r'[\x00-\x08\x0b\x0c\x0e-\x1f]', '', s)


def excel_workbook_bytes(
    feuilles: Sequence[Tuple[str, Sequence[str], Sequence[Sequence[Any]]]],
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True)
    header_fill = PatternFill(fill_type='solid', start_color='FFDCE7F3', end_color='FFDCE7F3')
    align_wrap = Alignment(wrap_text=True, vertical='top')

    for idx, (title, headers, rows) in enumerate(feuilles):
        name = (title or 'Feuille')[:31] or f'S{idx + 1}'
        ws = wb.create_sheet(title=name)
        for c, h in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = align_wrap
        for r, row in enumerate(rows, start=2):
            for c, val in enumerate(row, start=1):
                v = val
                if isinstance(v, Decimal):
                    v = float(v)
                cell = ws.cell(row=r, column=c, value=v)
                cell.alignment = align_wrap
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def pdf_table_bytes(
    title: str,
    subtitle: str,
    headers: Sequence[str],
    rows: Iterable[Sequence[Any]],
    *,
    landscape: bool = False,
    font_size: int = 7,
) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape as rl_landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import cm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    def esc(t: Any) -> str:
        if t is None:
            return ''
        s = str(t)
        return s.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')

    rows_list: List[List[str]] = [[cell_str(x) for x in row] for row in rows]
    hdr = [esc(h) for h in headers]
    body = []
    sty = getSampleStyleSheet()
    hdr_para = [
        Paragraph(f'<font size="{font_size}"><b>{h}</b></font>', sty['Normal'])
        for h in hdr
    ]
    for row in rows_list:
        body.append(
            [
                Paragraph(f'<font size="{font_size}">{esc(c)}</font>', sty['Normal'])
                for c in row
            ]
        )

    buf = io.BytesIO()
    page = rl_landscape(A4) if landscape else A4
    doc = SimpleDocTemplate(
        buf,
        pagesize=page,
        leftMargin=0.8 * cm,
        rightMargin=0.8 * cm,
        topMargin=0.9 * cm,
        bottomMargin=0.9 * cm,
    )
    story = []
    story.append(Paragraph(f'<b>{esc(title)}</b>', sty['Title']))
    if subtitle:
        story.append(Paragraph(f'<font size="9">{esc(subtitle)}</font>', sty['Normal']))
    story.append(Spacer(1, 0.35 * cm))

    table_data = [hdr_para] + body
    t = Table(table_data, repeatRows=1)
    t.setStyle(
        TableStyle(
            [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E2E8F0')),
                ('VALIGN', (0, 0), (-1, -1), 'TOP'),
                ('LEFTPADDING', (0, 0), (-1, -1), 3),
                ('RIGHTPADDING', (0, 0), (-1, -1), 3),
                ('GRID', (0, 0), (-1, -1), 0.25, colors.grey),
            ]
        )
    )
    story.append(t)
    doc.build(story)
    return buf.getvalue()


def response_attachment_xlsx(content: bytes, filename: str) -> HttpResponse:
    r = HttpResponse(
        content,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    r['Content-Disposition'] = f'attachment; filename="{filename}"'
    return r


def response_attachment_pdf(content: bytes, filename: str) -> HttpResponse:
    r = HttpResponse(content, content_type='application/pdf')
    r['Content-Disposition'] = f'attachment; filename="{filename}"'
    return r
