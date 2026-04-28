import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.urls import reverse
from django.views.decorators.http import require_POST
from django.core.paginator import Paginator
from django.db.models import Q, Count

from entreprise.models import Categorie, SousCategorie, Produit, Entreprise
from utilities.utility import error_message_list
from .forms import CategorieForm, SousCategorieForm, ProduitForm, ImportExcelForm


def _get_entreprises(user):
    return Entreprise.objects.filter(user=user)


# ── CATÉGORIES ────────────────────────────────────────────────────────────────

@login_required
def liste_categories(request):
    categories = (
        Categorie.objects.filter(entreprise__user=request.user)
        .select_related('entreprise')
        .annotate(
            nb_sous=Count('sous_categories', distinct=True),
            nb_produits=Count('sous_categories__produits', distinct=True),
        )
        .order_by('entreprise__nom', 'nom')
    )
    form = CategorieForm(user=request.user)
    context = {
        'categories': categories,
        'form': form,
        'cat_actif': True,
        'produit_actif': True,
    }
    if request.headers.get('HX-Request') and request.headers.get('HX-Target') == 'cat-contenu':
        return render(request, 'produit/categories/partial/contenu.html', context)
    return render(request, 'produit/categories/liste.html', context)


@login_required
def creer_categorie(request):
    if request.method == 'POST':
        form = CategorieForm(request.POST, user=request.user)
        if form.is_valid():
            cat = form.save()
            messages.success(request, f"Catégorie « {cat.nom} » créée.")
            response = HttpResponse(status=204)
            response['HX-Redirect'] = reverse('produit:categories')
            return response
        messages.info(request, error_message_list(form))
        categories = (
            Categorie.objects.filter(entreprise__user=request.user)
            .select_related('entreprise')
            .annotate(nb_sous=Count('sous_categories', distinct=True), nb_produits=Count('sous_categories__produits', distinct=True))
            .order_by('entreprise__nom', 'nom')
        )
        return render(request, 'produit/categories/partial/contenu.html', {
            'categories': categories, 'form': form, 'cat_actif': True, 'produit_actif': True,
        })
    return redirect('produit:categories')


@login_required
def modifier_categorie(request, pk):
    categorie = get_object_or_404(Categorie, pk=pk, entreprise__user=request.user)
    form = CategorieForm(instance=categorie, user=request.user)
    context = {'form': form, 'categorie': categorie, 'cat_actif': True, 'produit_actif': True}

    if request.method == 'POST':
        form = CategorieForm(request.POST, instance=categorie, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, f"Catégorie « {categorie.nom} » modifiée.")
            response = HttpResponse(status=204)
            response['HX-Redirect'] = reverse('produit:categories')
            return response
        messages.info(request, error_message_list(form))
        return render(request, 'produit/categories/partial/form_mod.html', context)

    if request.headers.get('HX-Request'):
        return render(request, 'produit/categories/partial/form_mod.html', context)
    return redirect('produit:categories')


@require_POST
@login_required
def supprimer_categorie(request, pk):
    categorie = get_object_or_404(Categorie, pk=pk, entreprise__user=request.user)
    nom = categorie.nom
    try:
        categorie.delete()
        messages.success(request, f"Catégorie « {nom} » supprimée.")
    except Exception:
        messages.info(request, f"Impossible de supprimer « {nom} » — des produits y sont liés.")
    response = HttpResponse(status=204)
    response['HX-Redirect'] = reverse('produit:categories')
    return response


# ── SOUS-CATÉGORIES ───────────────────────────────────────────────────────────

@login_required
def liste_sous_categories(request):
    q = request.GET.get('q', '').strip()
    cat_filter = request.GET.get('cat', '')

    qs = (
        SousCategorie.objects.filter(categorie__entreprise__user=request.user)
        .select_related('categorie', 'categorie__entreprise')
        .annotate(nb_produits=Count('produits', distinct=True))
        .order_by('categorie__entreprise__nom', 'categorie__nom', 'nom')
    )
    if q:
        qs = qs.filter(Q(nom__icontains=q) | Q(categorie__nom__icontains=q))
    if cat_filter:
        qs = qs.filter(categorie_id=cat_filter)

    categories = Categorie.objects.filter(entreprise__user=request.user).select_related('entreprise').order_by('entreprise__nom', 'nom')
    form = SousCategorieForm(user=request.user)
    context = {
        'sous_categories': qs,
        'categories': categories,
        'form': form,
        'q': q,
        'cat_filter': cat_filter,
        'souscat_actif': True,
        'produit_actif': True,
    }
    if request.headers.get('HX-Request') and request.headers.get('HX-Target') == 'souscat-contenu':
        return render(request, 'produit/sous-categories/partial/contenu.html', context)
    return render(request, 'produit/sous-categories/liste.html', context)


@login_required
def creer_sous_categorie(request):
    if request.method == 'POST':
        form = SousCategorieForm(request.POST, user=request.user)
        if form.is_valid():
            sc = form.save()
            messages.success(request, f"Sous-catégorie « {sc.nom} » créée.")
            response = HttpResponse(status=204)
            response['HX-Redirect'] = reverse('produit:sous-categories')
            return response
        messages.info(request, error_message_list(form))
        qs = (
            SousCategorie.objects.filter(categorie__entreprise__user=request.user)
            .select_related('categorie', 'categorie__entreprise')
            .annotate(nb_produits=Count('produits', distinct=True))
        )
        categories = Categorie.objects.filter(entreprise__user=request.user).select_related('entreprise')
        return render(request, 'produit/sous-categories/partial/contenu.html', {
            'sous_categories': qs, 'categories': categories, 'form': form,
            'souscat_actif': True, 'produit_actif': True,
        })
    return redirect('produit:sous-categories')


@login_required
def modifier_sous_categorie(request, pk):
    sc = get_object_or_404(SousCategorie, pk=pk, categorie__entreprise__user=request.user)
    form = SousCategorieForm(instance=sc, user=request.user)
    context = {'form': form, 'sous_categorie': sc, 'souscat_actif': True, 'produit_actif': True}

    if request.method == 'POST':
        form = SousCategorieForm(request.POST, instance=sc, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, f"Sous-catégorie « {sc.nom} » modifiée.")
            response = HttpResponse(status=204)
            response['HX-Redirect'] = reverse('produit:sous-categories')
            return response
        messages.info(request, error_message_list(form))
        return render(request, 'produit/sous-categories/partial/form_mod.html', context)

    if request.headers.get('HX-Request'):
        return render(request, 'produit/sous-categories/partial/form_mod.html', context)
    return redirect('produit:sous-categories')


@require_POST
@login_required
def supprimer_sous_categorie(request, pk):
    sc = get_object_or_404(SousCategorie, pk=pk, categorie__entreprise__user=request.user)
    nom = sc.nom
    try:
        sc.delete()
        messages.success(request, f"Sous-catégorie « {nom} » supprimée.")
    except Exception:
        messages.info(request, f"Impossible de supprimer « {nom} » — des produits y sont liés.")
    response = HttpResponse(status=204)
    response['HX-Redirect'] = reverse('produit:sous-categories')
    return response


# ── PRODUITS ──────────────────────────────────────────────────────────────────

@login_required
def liste_produits(request):
    q = request.GET.get('q', '').strip()
    cat_id = request.GET.get('cat', '')
    scat_id = request.GET.get('scat', '')
    statut = request.GET.get('statut', '')

    qs = (
        Produit.objects.filter(sous_categorie__categorie__entreprise__user=request.user)
        .select_related(
            'entreprise',
            'sous_categorie',
            'sous_categorie__categorie',
            'sous_categorie__categorie__entreprise',
        )
        .order_by('sous_categorie__categorie__nom', 'sous_categorie__nom', 'nom')
    )
    if q:
        qs = qs.filter(
            Q(nom__icontains=q)
            | Q(code_barre__icontains=q)
            | Q(sku__icontains=q)
            | Q(description__icontains=q)
        )
    if cat_id:
        qs = qs.filter(sous_categorie__categorie_id=cat_id)
    if scat_id:
        qs = qs.filter(sous_categorie_id=scat_id)
    if statut == '1':
        qs = qs.filter(est_actif=True)
    elif statut == '0':
        qs = qs.filter(est_actif=False)

    paginator = Paginator(qs, 20)
    page_obj = paginator.get_page(request.GET.get('page', 1))

    categories = Categorie.objects.filter(entreprise__user=request.user).select_related('entreprise').order_by('nom')
    sous_cats = SousCategorie.objects.filter(categorie__entreprise__user=request.user).select_related('categorie')
    if cat_id:
        sous_cats = sous_cats.filter(categorie_id=cat_id)

    context = {
        'produits': page_obj,
        'categories': categories,
        'sous_cats': sous_cats,
        'q': q, 'cat_id': cat_id, 'scat_id': scat_id, 'statut': statut,
        'produit_liste_actif': True,
        'produit_actif': True,
    }
    if request.headers.get('HX-Request') and request.headers.get('HX-Target') == 'produits-table':
        return render(request, 'produit/produits/partial/table.html', context)
    return render(request, 'produit/produits/liste.html', context)


@login_required
def detail_produit(request, pk):
    produit = get_object_or_404(
        Produit, pk=pk, sous_categorie__categorie__entreprise__user=request.user
    )
    stocks = produit.niveaux_stock.select_related('depot', 'pointdevente').all()

    # Calculs de prix/marge
    marge_brute = produit.prix_vente_ht - produit.prix_achat_ht
    prix_ttc = produit.prix_vente_ttc
    montant_tva = prix_ttc - produit.prix_vente_ht
    taux_marge = (
        (marge_brute / produit.prix_achat_ht * 100)
        if produit.prix_achat_ht > 0 else None
    )

    return render(request, 'produit/produits/detail.html', {
        'produit': produit,
        'stocks': stocks,
        'marge_brute': marge_brute,
        'montant_tva': montant_tva,
        'taux_marge': taux_marge,
        'produit_liste_actif': True,
        'produit_actif': True,
    })


@login_required
def creer_produit(request):
    form = ProduitForm(user=request.user)
    context = {'form': form, 'produit_liste_actif': True, 'produit_actif': True}

    if request.method == 'POST':
        form = ProduitForm(request.POST, request.FILES, user=request.user)
        if form.is_valid():
            produit = form.save()
            messages.success(request, f"Produit « {produit.nom} » créé avec succès.")
            return redirect('produit:detail', pk=produit.pk)
        messages.info(request, error_message_list(form))
        context['form'] = form
    return render(request, 'produit/produits/form.html', context)


@login_required
def modifier_produit(request, pk):
    produit = get_object_or_404(
        Produit, pk=pk, sous_categorie__categorie__entreprise__user=request.user
    )
    form = ProduitForm(instance=produit, user=request.user)
    context = {'form': form, 'produit': produit, 'produit_liste_actif': True, 'produit_actif': True}

    if request.method == 'POST':
        form = ProduitForm(request.POST, request.FILES, instance=produit, user=request.user)
        if form.is_valid():
            form.save()
            messages.success(request, f"Produit « {produit.nom} » modifié.")
            return redirect('produit:detail', pk=produit.pk)
        messages.info(request, error_message_list(form))
        context['form'] = form
    return render(request, 'produit/produits/form.html', context)


@require_POST
@login_required
def toggle_produit(request, pk):
    produit = get_object_or_404(
        Produit, pk=pk, sous_categorie__categorie__entreprise__user=request.user
    )
    produit.est_actif = not produit.est_actif
    produit.save(update_fields=['est_actif'])
    etat = "activé" if produit.est_actif else "désactivé"
    messages.success(request, f"Produit « {produit.nom} » {etat}.")
    response = HttpResponse(status=204)
    response['HX-Redirect'] = reverse('produit:liste')
    return response


@require_POST
@login_required
def supprimer_produit(request, pk):
    produit = get_object_or_404(
        Produit, pk=pk, sous_categorie__categorie__entreprise__user=request.user
    )
    nom = produit.nom
    try:
        produit.delete()
        messages.success(request, f"Produit « {nom} » supprimé.")
    except Exception:
        messages.info(request, f"Impossible de supprimer « {nom} » — il est utilisé dans des documents.")
    response = HttpResponse(status=204)
    response['HX-Redirect'] = reverse('produit:liste')
    return response


# ── IMPORT EXCEL ──────────────────────────────────────────────────────────────

@login_required
def import_produits(request):
    form = ImportExcelForm()
    results = None
    context = {'form': form, 'import_actif': True, 'produit_actif': True}

    if request.method == 'POST':
        form = ImportExcelForm(request.POST, request.FILES)
        if form.is_valid():
            fichier = form.cleaned_data['fichier']
            try:
                wb = openpyxl.load_workbook(fichier, data_only=True)
                ws = wb.active
                headers = [
                    str(c.value).strip().lower().replace(' ', '_') if c.value else ''
                    for c in ws[1]
                ]

                created = updated = skipped = 0
                errors = []

                for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    data = dict(zip(headers, row))
                    nom = str(data.get('nom') or '').strip()
                    if not nom:
                        skipped += 1
                        continue

                    cat_nom = str(data.get('categorie') or '').strip()
                    scat_nom = str(data.get('sous_categorie') or '').strip()

                    if not cat_nom or not scat_nom:
                        errors.append(f"Ligne {row_idx} ({nom}): catégorie ou sous-catégorie manquante.")
                        continue

                    # Récupère l'entreprise depuis la colonne du fichier
                    entreprise_nom = str(data.get('entreprise') or '').strip()
                    if not entreprise_nom:
                        errors.append(f"Ligne {row_idx} ({nom}): colonne 'entreprise' manquante ou vide.")
                        continue
                    entreprise = Entreprise.objects.filter(nom__iexact=entreprise_nom).first()
                    if not entreprise:
                        errors.append(f"Ligne {row_idx} ({nom}): entreprise « {entreprise_nom} » introuvable dans le système.")
                        continue

                    cat, _ = Categorie.objects.get_or_create(entreprise=entreprise, nom=cat_nom)
                    scat, _ = SousCategorie.objects.get_or_create(categorie=cat, nom=scat_nom)

                    def to_dec(val, default=0):
                        try:
                            return float(val) if val not in (None, '') else default
                        except (ValueError, TypeError):
                            return default

                    unite = str(data.get('unite_mesure') or 'PCS').strip().upper()[:10]
                    methode = str(data.get('methode_gestion') or 'FEFO').strip().upper()[:30]
                    valides_unite = [u[0] for u in Produit.UNITES]
                    valides_methode = [m[0] for m in Produit.METHODES]
                    if unite not in valides_unite:
                        unite = 'PCS'
                    if methode not in valides_methode:
                        methode = 'FEFO'

                    sku_raw = str(data.get('sku') or '').strip()[:100]
                    sku_val = sku_raw or None

                    defaults = {
                        'entreprise': entreprise,
                        'sous_categorie': scat,
                        'nom': nom,
                        'sku': sku_val,
                        'description': str(data.get('description') or '').strip(),
                        'prix_achat_ht': to_dec(data.get('prix_achat_ht'), 0),
                        'prix_vente_ht': to_dec(data.get('prix_vente_ht'), 0),
                        'tva_taux': to_dec(data.get('tva_taux'), 16),
                        'unite_mesure': unite,
                        'stock_alerte': to_dec(data.get('stock_alerte'), 5),
                        'methode_gestion': methode,
                        'vie': int(to_dec(data.get('vie'), 30)),
                    }

                    code_barre = str(data.get('code_barre') or '').strip() or None

                    try:
                        if code_barre:
                            _, is_new = Produit.objects.update_or_create(
                                code_barre=code_barre, defaults=defaults
                            )
                        elif sku_val:
                            _, is_new = Produit.objects.update_or_create(
                                sous_categorie=scat, sku=sku_val, defaults=defaults
                            )
                        else:
                            Produit.objects.create(code_barre=None, **defaults)
                            is_new = True

                        if is_new:
                            created += 1
                        else:
                            updated += 1
                    except Exception as e:
                        errors.append(f"Ligne {row_idx} ({nom}): {e}")

                results = {'created': created, 'updated': updated, 'skipped': skipped, 'errors': errors}
                if created or updated:
                    messages.success(request, f"{created} produit(s) créé(s), {updated} mis à jour.")

            except Exception as e:
                messages.info(request, f"Erreur de lecture du fichier : {e}")

        context['form'] = form
        context['results'] = results

    return render(request, 'produit/import/import.html', context)


@login_required
def telecharger_modele_excel(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Produits"

    headers = [
        'entreprise', 'nom', 'categorie', 'sous_categorie', 'code_barre', 'sku', 'description',
        'prix_achat_ht', 'prix_vente_ht', 'tva_taux', 'unite_mesure',
        'stock_alerte', 'methode_gestion', 'vie',
    ]

    h_font = Font(bold=True, color='FFFFFF')
    h_fill = PatternFill('solid', fgColor='1A56DB')
    h_align = Alignment(horizontal='center', vertical='center')
    # Colonne entreprise mise en évidence (orange)
    e_fill = PatternFill('solid', fgColor='E07B00')

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = h_font
        cell.fill = e_fill if h == 'entreprise' else h_fill
        cell.alignment = h_align
        ws.column_dimensions[cell.column_letter].width = 24 if h == 'entreprise' else 20

    ws.row_dimensions[1].height = 22

    # Exemples de données — les noms d'entreprise doivent correspondre exactement
    ws.append([
        'Mon Entreprise SA', 'Savon Lux', 'Hygiène', 'Soins du corps', '6901234567890', 'HYG-SAV-LUX',
        'Savon de toilette 150g', 1500, 2500, 16, 'PCS', 10, 'FEFO', 365,
    ])
    ws.append([
        'Mon Entreprise SA', 'Sucre 1kg', 'Alimentation', 'Épicerie', '', 'ALIM-SUC-1KG',
        '', 800, 1200, 16, 'KG', 20, 'FIFO', 730,
    ])

    # Feuille d'instructions
    ws2 = wb.create_sheet("Instructions")
    ws2.column_dimensions['A'].width = 22
    ws2.column_dimensions['B'].width = 12
    ws2.column_dimensions['C'].width = 65
    ws2.append(['Colonne', 'Obligatoire', 'Description / Valeurs acceptées'])
    ws2['A1'].font = Font(bold=True)
    ws2['B1'].font = Font(bold=True)
    ws2['C1'].font = Font(bold=True)
    for row in [
        ('entreprise', 'OUI', 'Nom exact de l\'entreprise dans le système (sensible à la casse ignorée)'),
        ('nom', 'OUI', 'Nom du produit'),
        ('categorie', 'OUI', 'Nom de catégorie (créée automatiquement si nouvelle)'),
        ('sous_categorie', 'OUI', 'Nom de sous-catégorie (créée automatiquement si nouvelle)'),
        ('code_barre', 'NON', 'Code barre unique — laisser vide si aucun ; sert d’identifiant prioritaire pour mise à jour'),
        ('sku', 'NON', 'Référence SKU — unique par entreprise ; si pas de code-barre, sert à mettre à jour un produit existant'),
        ('description', 'NON', 'Description libre du produit'),
        ('prix_achat_ht', 'OUI', 'Prix achat hors taxe (nombre décimal)'),
        ('prix_vente_ht', 'OUI', 'Prix vente hors taxe (nombre décimal)'),
        ('tva_taux', 'NON', 'Taux TVA en % — défaut : 16'),
        ('unite_mesure', 'NON', 'PCS / KG / L / M / BOX — défaut : PCS'),
        ('stock_alerte', 'NON', 'Seuil alerte stock — défaut : 5'),
        ('methode_gestion', 'NON', 'FIFO / FEFO / LIFO — défaut : FEFO'),
        ('vie', 'NON', 'Durée de vie en jours — défaut : 30'),
    ]:
        ws2.append(row)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename="modele_import_produits.xlsx"'
    return response
